from collections.abc import Callable
from copy import copy
from datetime import date
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.worksheet.worksheet import Worksheet
from PIL import Image as PillowImage


class ContractWorkbookError(RuntimeError):
    pass


class ContractWorkbookRenderer:
    def __init__(
        self,
        *,
        template_path: Path,
        image_loader: Callable[[str], bytes | None] | None = None,
    ) -> None:
        self._template_path = template_path
        self._image_loader = image_loader or (lambda _object_key: None)

    def render(self, snapshot: dict[str, Any]) -> bytes:
        lines = list(snapshot.get("lines") or [])
        if not lines:
            raise ContractWorkbookError("contract requires at least one line")
        workbook = load_workbook(self._template_path)
        if workbook.sheetnames != ["合同"]:
            raise ContractWorkbookError("contract template sheets are invalid")
        sheet = workbook["合同"]
        extra_rows = self._prepare_detail_area(sheet, len(lines))
        self._write_header(sheet, snapshot)
        self._write_lines(sheet, lines, snapshot)
        self._write_totals(sheet, len(lines), extra_rows)
        self._write_delivery_term(sheet, snapshot, extra_rows)
        self._write_supplier_signature(sheet, snapshot, extra_rows)
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
        workbook.calculation.calcMode = "auto"
        output = BytesIO()
        workbook.save(output)
        return output.getvalue()

    @staticmethod
    def _prepare_detail_area(sheet: Worksheet, line_count: int) -> int:
        product_column_styles = [
            copy(sheet.cell(8, column)._style)  # type: ignore[union-attr]
            for column in range(1, 4)
        ]
        for merged_range in list(sheet.merged_cells.ranges):
            if merged_range.min_row <= 19 and merged_range.max_row >= 8:
                sheet.unmerge_cells(str(merged_range))
        extra_rows = max(line_count - 12, 0)
        if extra_rows:
            shifted_merges = [
                (
                    str(merged_range),
                    merged_range.min_row,
                    merged_range.min_col,
                    merged_range.max_row,
                    merged_range.max_col,
                )
                for merged_range in list(sheet.merged_cells.ranges)
                if merged_range.min_row >= 20
            ]
            for address, *_coordinates in shifted_merges:
                sheet.unmerge_cells(address)
            sheet.insert_rows(20, extra_rows)
            source_height = sheet.row_dimensions[19].height
            for row_index in range(20, 20 + extra_rows):
                sheet.row_dimensions[row_index].height = source_height
                for column in range(1, 10):
                    source = sheet.cell(19, column)
                    target = sheet.cell(row_index, column)
                    target._style = copy(source._style)  # type: ignore[union-attr]
                    target.number_format = source.number_format
                    target.alignment = copy(source.alignment)  # type: ignore[assignment]
                    target.protection = copy(source.protection)  # type: ignore[assignment]
            for _address, min_row, min_col, max_row, max_col in shifted_merges:
                sheet.merge_cells(
                    start_row=min_row + extra_rows,
                    start_column=min_col,
                    end_row=max_row + extra_rows,
                    end_column=max_col,
                )
            sheet.print_area = f"A1:I{51 + extra_rows}"
        detail_end = 19 + extra_rows
        for row_index in range(8, detail_end + 1):
            for column in range(1, 4):
                cell = sheet.cell(row_index, column)
                cell._style = copy(  # type: ignore[union-attr]
                    product_column_styles[column - 1]
                )
                border = copy(cell.border)
                header_border = sheet.cell(7, column).border
                border.left = copy(header_border.left)
                border.right = copy(header_border.right)
                border.bottom = copy(header_border.bottom)
                cell.border = border  # type: ignore[assignment]
        for row in sheet.iter_rows(min_row=8, max_row=detail_end, min_col=1, max_col=9):
            for cell in row:
                cell.value = None
        return extra_rows

    @staticmethod
    def _write_header(sheet: Worksheet, snapshot: dict[str, Any]) -> None:
        contract_no = str(snapshot["contractNo"])
        signing_date = date.fromisoformat(str(snapshot["signingDate"]))
        factory = dict(snapshot["factory"])
        sheet["G3"] = f"合同编号：{contract_no}"
        sheet["G4"] = (
            f"签订时间：{signing_date.year}年{signing_date.month}月{signing_date.day}日"
        )
        sheet["A4"] = f"供方：{factory['legalName']}"

    def _write_lines(
        self,
        sheet: Worksheet,
        lines: list[dict[str, Any]],
        snapshot: dict[str, Any],
    ) -> None:
        start_row = 8
        ship_date = date.fromisoformat(str(snapshot["contractShipDate"]))
        for offset, line in enumerate(lines):
            row = start_row + offset
            sheet.cell(row, 4, str(line["propertiesValue"]))
            sheet.cell(row, 5, int(line["quantity"]))
            sheet.cell(row, 6, None)
            sheet.cell(row, 7, f'=IF(OR(E{row}="",F{row}=""),"",E{row}*F{row})')
        group_start = start_row
        for index in range(1, len(lines) + 1):
            boundary = (
                index == len(lines)
                or lines[index]["productId"] != lines[index - 1]["productId"]
            )
            if not boundary:
                continue
            group_end = start_row + index - 1
            line = lines[index - 1]
            sheet.cell(group_start, 1, str(line.get("itemNo") or ""))
            sheet.cell(group_start, 2, str(line["productName"]))
            if group_end > group_start:
                for column in (1, 2, 3):
                    sheet.merge_cells(
                        start_row=group_start,
                        start_column=column,
                        end_row=group_end,
                        end_column=column,
                    )
            self._add_product_image(
                sheet,
                object_key=line.get("imageObjectKey"),
                start_row=group_start,
                end_row=group_end,
            )
            group_start = group_end + 1
        last_row = start_row + len(lines) - 1
        sheet["H8"] = f"{ship_date.year}年{ship_date.month}月{ship_date.day}日"
        detail_end = max(19, last_row)
        sheet.merge_cells(start_row=8, start_column=8, end_row=detail_end, end_column=8)
        sheet.merge_cells(start_row=8, start_column=9, end_row=detail_end, end_column=9)

    def _add_product_image(
        self,
        sheet: Worksheet,
        *,
        object_key: object,
        start_row: int,
        end_row: int,
    ) -> None:
        if not object_key:
            return
        try:
            content = self._image_loader(str(object_key))
            if not content or len(content) > 5 * 1024 * 1024:
                return
            with PillowImage.open(BytesIO(content)) as source:
                source.verify()
                width, height = source.size
            if width <= 0 or height <= 0 or width > 6000 or height > 6000:
                return
            max_width = 145.0
            max_height = sum(
                float(sheet.row_dimensions[row].height or 15) * 4 / 3
                for row in range(start_row, end_row + 1)
            )
            scale = min(max_width / width, max_height / height, 1.0)
            image = OpenpyxlImage(BytesIO(content))
            image.width = width * scale
            image.height = height * scale
            sheet.add_image(image, f"C{start_row}")
        except Exception:
            return

    @staticmethod
    def _write_totals(sheet: Worksheet, line_count: int, extra_rows: int) -> None:
        last_row = 8 + line_count - 1
        total_row = 20 + extra_rows
        words_row = total_row + 1
        sheet.cell(total_row, 5, f"=SUM(E8:E{last_row})")
        sheet.cell(total_row, 7, (
            f'=IF(COUNT(F8:F{last_row})=ROWS(F8:F{last_row}),SUM(G8:G{last_row}),"")'
        ))
        sheet.cell(words_row, 2, f'=IF(G{total_row}="","",G{total_row})')
        sheet.cell(words_row, 4, f'=IF(G{total_row}="","",G{total_row})')

    @staticmethod
    def _write_delivery_term(
        sheet: Worksheet, snapshot: dict[str, Any], extra_rows: int
    ) -> None:
        ship_date = date.fromisoformat(str(snapshot["contractShipDate"]))
        sheet.cell(24 + extra_rows, 1, (
            f"一.交货期限：{ship_date.year}年{ship_date.month}月{ship_date.day}日前全部出货"
        ))

    @staticmethod
    def _write_supplier_signature(
        sheet: Worksheet, snapshot: dict[str, Any], extra_rows: int
    ) -> None:
        factory = dict(snapshot["factory"])
        phone = str(factory.get("phone") or "")
        sheet.cell(45 + extra_rows, 5, (
            "                         供      方\n"
            f"单位名称（章）：{factory['legalName']}\n"
            f"单位地址：{factory['address']}\n"
            f"法定代表人：{factory['legalRepresentative']}\n"
            "委托代理人：\n"
            "开户银行：\n"
            "账号：\n"
            f"电话：{phone}"
        ))
