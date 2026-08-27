from dataclasses import dataclass
from io import BytesIO
from pathlib import PurePosixPath
from xml.etree import ElementTree
from zipfile import ZipFile

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

InspectionWorkbookIssue = dict[str, str | int]


class InspectionWorkbookValidationError(ValueError):
    def __init__(self, issues: tuple[InspectionWorkbookIssue, ...]) -> None:
        self.issues = issues
        super().__init__("inspection workbook validation failed")


@dataclass(frozen=True)
class InspectionWorkbookLimits:
    max_source_bytes: int = 20 * 1024 * 1024
    max_zip_entries: int = 20_000
    max_uncompressed_bytes: int = 100 * 1024 * 1024
    max_worksheets: int = 10
    max_data_rows: int = 5_000


@dataclass(frozen=True)
class InspectionWorkbookLine:
    source_row: int
    supplier_number: str
    factory_name: str
    source_sku_id: str
    source_product_id: str
    product_name: str
    properties_value: str
    quantity: int
    box_number: str
    reason: str | None


@dataclass(frozen=True)
class InspectionWorkbookSnapshot:
    supplier_number: str
    factory_name: str
    total_quantity: int
    box_numbers: tuple[str, ...]
    lines: tuple[InspectionWorkbookLine, ...]


class InspectionWorkbookParser:
    _BASE_HEADERS = (
        "编号",
        "厂家名称",
        "商品编码",
        "款式编码",
        "商品名称",
        "颜色/规格",
        "数量",
        "箱数",
    )

    def __init__(self, *, limits: InspectionWorkbookLimits | None = None) -> None:
        self._limits = limits or InspectionWorkbookLimits()

    def parse(self, content: bytes) -> InspectionWorkbookSnapshot:
        if len(content) > self._limits.max_source_bytes:
            raise InspectionWorkbookValidationError(
                (
                    {
                        "code": "file_too_large",
                        "message": "质检 Excel 超过文件大小上限",
                    },
                )
            )
        self._validate_archive(content)
        workbook = load_workbook(BytesIO(content), data_only=False, read_only=False)
        for extra_sheet in workbook.worksheets:
            if extra_sheet.title == "Sheet1":
                continue
            for cells in extra_sheet.iter_rows():
                for cell in cells:
                    if cell.value is not None:
                        raise InspectionWorkbookValidationError(
                            (
                                {
                                    "code": "unexpected_sheet_data",
                                    "message": "Sheet1 以外的工作表不能包含业务数据",
                                    "sheet": extra_sheet.title,
                                    "row": cell.row,
                                    "field": cell.column_letter,
                                },
                            )
                        )
        sheet = workbook["Sheet1"]
        if sheet.max_row - 1 > self._limits.max_data_rows:
            raise InspectionWorkbookValidationError(
                (
                    {
                        "code": "too_many_data_rows",
                        "message": "质检 Excel 数据行数超过上限",
                        "sheet": "Sheet1",
                    },
                )
            )
        headers = tuple(sheet.cell(1, column).value for column in range(1, 9))
        for column, (actual, expected) in enumerate(
            zip(headers, self._BASE_HEADERS, strict=True), 1
        ):
            if actual != expected:
                field = get_column_letter(column)
                raise InspectionWorkbookValidationError(
                    (
                        {
                            "code": "invalid_header",
                            "message": f"{field} 列表头应为“{expected}”",
                            "sheet": "Sheet1",
                            "row": 1,
                            "field": field,
                        },
                    )
                )

        lines: list[InspectionWorkbookLine] = []
        for row in range(2, sheet.max_row + 1):
            values = [sheet.cell(row, column).value for column in range(1, 8)]
            if all(value is None for value in values):
                for later_row in range(row + 1, sheet.max_row + 1):
                    later_values = [sheet.cell(later_row, column).value for column in range(1, 8)]
                    if any(value is not None for value in later_values):
                        raise InspectionWorkbookValidationError(
                            (
                                {
                                    "code": "data_after_blank_row",
                                    "message": "空白数据行之后不能再出现业务数据",
                                    "sheet": "Sheet1",
                                    "row": later_row,
                                },
                            )
                        )
                break
            for column, value in enumerate(values, 1):
                if value is None or (isinstance(value, str) and not value.strip()):
                    field = get_column_letter(column)
                    raise InspectionWorkbookValidationError(
                        (
                            {
                                "code": "required_field_missing",
                                "message": f"{self._BASE_HEADERS[column - 1]}不能为空",
                                "sheet": "Sheet1",
                                "row": row,
                                "field": field,
                            },
                        )
                    )
            quantity = values[6]
            if not isinstance(quantity, int) or isinstance(quantity, bool) or quantity <= 0:
                raise InspectionWorkbookValidationError(
                    (
                        {
                            "code": "invalid_quantity",
                            "message": "仓库退回数量必须是正整数",
                            "sheet": "Sheet1",
                            "row": row,
                            "field": "G",
                        },
                    )
                )
            box_number = self._merged_value(sheet, row=row, column=8)
            if box_number is None or (isinstance(box_number, str) and not box_number.strip()):
                raise InspectionWorkbookValidationError(
                    (
                        {
                            "code": "required_field_missing",
                            "message": "箱数不能为空",
                            "sheet": "Sheet1",
                            "row": row,
                            "field": "H",
                        },
                    )
                )
            reason_value = self._merged_value(sheet, row=row, column=9)
            lines.append(
                InspectionWorkbookLine(
                    source_row=row,
                    supplier_number=str(values[0]),
                    factory_name=str(values[1]),
                    source_sku_id=str(values[2]),
                    source_product_id=str(values[3]),
                    product_name=str(values[4]),
                    properties_value=str(values[5]),
                    quantity=quantity,
                    box_number=str(box_number),
                    reason=str(reason_value) if reason_value is not None else None,
                )
            )

        box_numbers = tuple(dict.fromkeys(line.box_number for line in lines))
        return InspectionWorkbookSnapshot(
            supplier_number=lines[0].supplier_number,
            factory_name=lines[0].factory_name,
            total_quantity=sum(line.quantity for line in lines),
            box_numbers=box_numbers,
            lines=tuple(lines),
        )

    def _validate_archive(self, content: bytes) -> None:
        with ZipFile(BytesIO(content)) as archive:
            entries = archive.infolist()
            if len(entries) > self._limits.max_zip_entries:
                raise InspectionWorkbookValidationError(
                    (
                        {
                            "code": "too_many_zip_entries",
                            "message": "质检 Excel 内部文件数超过上限",
                        },
                    )
                )
            if any(
                PurePosixPath(entry.filename).is_absolute()
                or ".." in PurePosixPath(entry.filename).parts
                or "\\" in entry.filename
                for entry in entries
            ):
                raise InspectionWorkbookValidationError(
                    (
                        {
                            "code": "unsafe_archive_path",
                            "message": "质检 Excel 包含不安全的内部路径",
                        },
                    )
                )
            worksheet_count = sum(
                entry.filename.startswith("xl/worksheets/") and entry.filename.endswith(".xml")
                for entry in entries
            )
            if worksheet_count > self._limits.max_worksheets:
                raise InspectionWorkbookValidationError(
                    (
                        {
                            "code": "too_many_worksheets",
                            "message": "质检 Excel 工作表数量超过上限",
                        },
                    )
                )
            if sum(entry.file_size for entry in entries) > self._limits.max_uncompressed_bytes:
                raise InspectionWorkbookValidationError(
                    (
                        {
                            "code": "uncompressed_content_too_large",
                            "message": "质检 Excel 解压后总量超过上限",
                        },
                    )
                )
            for entry in entries:
                if not entry.filename.endswith(".rels"):
                    continue
                relationships = ElementTree.fromstring(archive.read(entry))
                if any(
                    relationship.attrib.get("TargetMode") == "External"
                    for relationship in relationships
                ):
                    raise InspectionWorkbookValidationError(
                        (
                            {
                                "code": "external_relationship",
                                "message": "质检 Excel 不允许引用外部资源",
                            },
                        )
                    )

    @staticmethod
    def _merged_value(sheet: Worksheet, *, row: int, column: int) -> object:
        value = sheet.cell(row, column).value
        if value is not None:
            return value
        coordinate = sheet.cell(row, column).coordinate
        for merged_range in sheet.merged_cells.ranges:
            if coordinate in merged_range:
                return sheet.cell(merged_range.min_row, merged_range.min_col).value
        return None
