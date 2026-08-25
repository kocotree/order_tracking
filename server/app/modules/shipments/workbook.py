from collections import defaultdict
from copy import copy
from dataclasses import dataclass
from datetime import date
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


class ShipmentWorkbookError(RuntimeError):
    pass


@dataclass(frozen=True)
class ShipmentWorkbookLine:
    order_no: str
    box_no: str
    sku_id: str
    product_name: str
    properties_value: str
    packed_quantity: int
    total_quantity: int


@dataclass(frozen=True)
class ShipmentWorkbookSnapshot:
    business_date: date
    total_boxes: int
    lines: list[ShipmentWorkbookLine]


class ShipmentWorkbookRenderer:
    DETAIL_START_ROW = 3
    DETAIL_TEMPLATE_END_ROW = 23

    def __init__(self, *, template_path: Path) -> None:
        self._template_path = template_path

    def render(self, snapshot: ShipmentWorkbookSnapshot) -> bytes:
        if not snapshot.lines:
            raise ShipmentWorkbookError("shipment workbook requires detail lines")
        if snapshot.total_boxes <= 0:
            raise ShipmentWorkbookError("shipment workbook requires boxes")
        workbook = load_workbook(self._template_path)
        if workbook.sheetnames != ["发货明细", "汇总", "Sheet3"]:
            raise ShipmentWorkbookError("shipment template sheets are invalid")
        workbook.remove(workbook["Sheet3"])
        detail = workbook["发货明细"]
        summary = workbook["汇总"]
        self._prepare_detail_rows(detail, len(snapshot.lines))
        self._write_detail(detail, snapshot)
        self._write_summary(summary, snapshot)
        output = BytesIO()
        workbook.save(output)
        return output.getvalue()

    @classmethod
    def _prepare_detail_rows(cls, sheet: Worksheet, line_count: int) -> None:
        template_capacity = cls.DETAIL_TEMPLATE_END_ROW - cls.DETAIL_START_ROW + 1
        extra_rows = max(line_count - template_capacity, 0)
        if extra_rows:
            sheet.insert_rows(cls.DETAIL_TEMPLATE_END_ROW + 1, extra_rows)
            source_row = cls.DETAIL_TEMPLATE_END_ROW
            for row_index in range(source_row + 1, source_row + extra_rows + 1):
                sheet.row_dimensions[row_index].height = sheet.row_dimensions[source_row].height
                for column in range(1, 8):
                    source = sheet.cell(source_row, column)
                    target = sheet.cell(row_index, column)
                    target._style = copy(source._style)  # type: ignore[union-attr]
                    target.number_format = source.number_format
                    target.alignment = copy(source.alignment)  # type: ignore[assignment]
                    target.protection = copy(source.protection)  # type: ignore[assignment]
        detail_end = cls.DETAIL_START_ROW + max(line_count, template_capacity) - 1
        for row in sheet.iter_rows(
            min_row=cls.DETAIL_START_ROW,
            max_row=detail_end,
            min_col=1,
            max_col=7,
        ):
            for cell in row:
                cell.value = None

    @classmethod
    def _write_detail(cls, sheet: Worksheet, snapshot: ShipmentWorkbookSnapshot) -> None:
        value = snapshot.business_date
        sheet["A1"] = (
            f"KK发货清单 {value.year}年{value.month}月{value.day}日 "
            f"共计{snapshot.total_boxes}箱"
        )
        for offset, line in enumerate(snapshot.lines):
            row = cls.DETAIL_START_ROW + offset
            values: tuple[str | int, ...] = (
                line.order_no,
                line.box_no,
                line.sku_id,
                line.product_name,
                line.properties_value,
                line.packed_quantity,
                line.total_quantity,
            )
            for column, cell_value in enumerate(values, 1):
                sheet.cell(row, column, cell_value)

    @staticmethod
    def _write_summary(sheet: Worksheet, snapshot: ShipmentWorkbookSnapshot) -> None:
        if sheet.max_row > 1:
            sheet.delete_rows(2, sheet.max_row - 1)
        totals: dict[tuple[str, str], int] = defaultdict(int)
        for line in snapshot.lines:
            totals[(line.product_name, line.properties_value)] += line.total_quantity
        for row, ((product_name, properties_value), quantity) in enumerate(
            sorted(totals.items()), 2
        ):
            sheet.cell(row, 1, snapshot.business_date)
            sheet.cell(row, 1).number_format = "yyyy-mm-dd"
            sheet.cell(row, 2, product_name)
            sheet.cell(row, 3, properties_value)
            sheet.cell(row, 4, quantity)
