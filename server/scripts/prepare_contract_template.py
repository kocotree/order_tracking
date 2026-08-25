import argparse
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import PatternFill
from openpyxl.utils import column_index_from_string

PROJECT_ROOT = Path(__file__).resolve().parents[2]
TARGET = PROJECT_ROOT / "server/app/templates/processing_contract_v1.xlsx"


def prepare_template(source: Path, target: Path = TARGET) -> None:
    workbook = load_workbook(source)
    for sheet_name in list(workbook.sheetnames):
        if sheet_name != "合同":
            del workbook[sheet_name]
    sheet = workbook["合同"]
    for merged_range in list(sheet.merged_cells.ranges):
        if merged_range.min_row > 51 or merged_range.max_row > 51:
            sheet.unmerge_cells(str(merged_range))
    if sheet.max_row > 51:
        sheet.delete_rows(52, sheet.max_row - 51)
    if sheet.max_column > 9:
        sheet.delete_cols(10, sheet.max_column - 9)
    for key in list(sheet.column_dimensions):
        if column_index_from_string(key) > 9:
            del sheet.column_dimensions[key]
    for key in list(sheet.row_dimensions):
        if key > 51:
            del sheet.row_dimensions[key]
    sheet._images = []
    for row in sheet.iter_rows(min_row=8, max_row=19, min_col=1, max_col=9):
        for cell in row:
            if not isinstance(cell, MergedCell):
                cell.value = None
    sheet["G3"] = None
    sheet["G4"] = None
    sheet["A4"] = "供方："
    sheet["E20"] = None
    sheet["G20"] = None
    sheet["B21"] = None
    sheet["D21"] = None
    sheet["A24"] = None
    sheet["E45"] = (
        "                         供      方\n"
        "单位名称（章）：\n"
        "单位地址：\n"
        "法定代表人：\n"
        "委托代理人：\n"
        "开户银行：\n"
        "账号：\n"
        "电话："
    )
    for row in sheet.iter_rows(min_row=1, max_row=51, min_col=1, max_col=9):
        for cell in row:
            if not isinstance(cell, MergedCell):
                cell.fill = PatternFill()
    sheet.print_area = "A1:I51"
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"
    target.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(target)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source", type=Path, required=True)
    parser.add_argument("--target", type=Path, default=TARGET)
    args = parser.parse_args()
    prepare_template(args.source, args.target)


if __name__ == "__main__":
    main()
