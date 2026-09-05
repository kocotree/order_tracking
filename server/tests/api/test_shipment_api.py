import base64
import hashlib
from datetime import date, datetime
from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import Engine, delete, select
from sqlalchemy.orm import Session, sessionmaker

from app.adapters.private_files import FakePrivateFileStore
from app.db.models import (
    AuditLog,
    Factory,
    IdempotencyRecord,
    Order,
    OrderAssignment,
    OrderCompletionRecord,
    OrderLine,
    OutboxMessage,
    Product,
    ProductVariant,
    QuantityLedger,
    Shipment,
    ShipmentBox,
    ShipmentBoxItem,
    ShipmentLine,
    ShipmentNumberCounter,
    ShipmentReturnEvent,
    ShipmentReturnLine,
    ShipmentVoidRequest,
    StoredFile,
    User,
    UserSession,
)
from app.main import create_app
from app.modules.identity_access import IdentityAccessService

USER_IDS = ["shipment-api-factory-a-user", "shipment-api-factory-b-user"]
SAME_FACTORY_USER_ID = "shipment-api-factory-a-colleague"
ALL_FACTORY_USER_IDS = [*USER_IDS, SAME_FACTORY_USER_ID]
ADMIN_ID = "shipment-api-admin"
FACTORY_IDS = ["shipment-api-factory-a", "shipment-api-factory-b"]
ORDER_ID = "shipment-api-order-a"
PRODUCT_ID = "shipment-api-product"
VARIANT_ID = "shipment-api-variant"


def _clean(engine: Engine) -> None:
    with Session(engine) as session, session.begin():
        shipment_ids = select(Shipment.shipment_id).where(Shipment.factory_id.in_(FACTORY_IDS))
        box_ids = select(ShipmentBox.box_id).where(ShipmentBox.shipment_id.in_(shipment_ids))
        assignment_ids = select(OrderAssignment.order_assignment_id).where(
            OrderAssignment.factory_id.in_(FACTORY_IDS)
        )
        session.execute(
            delete(ShipmentVoidRequest).where(ShipmentVoidRequest.shipment_id.in_(shipment_ids))
        )
        return_event_ids = select(ShipmentReturnEvent.event_id).where(
            ShipmentReturnEvent.shipment_id.in_(shipment_ids)
        )
        session.execute(
            delete(ShipmentReturnLine).where(ShipmentReturnLine.event_id.in_(return_event_ids))
        )
        session.execute(
            delete(ShipmentReturnEvent).where(ShipmentReturnEvent.shipment_id.in_(shipment_ids))
        )
        session.execute(delete(QuantityLedger).where(QuantityLedger.source_id.in_(shipment_ids)))
        session.execute(delete(ShipmentLine).where(ShipmentLine.shipment_id.in_(shipment_ids)))
        session.execute(delete(ShipmentBoxItem).where(ShipmentBoxItem.box_id.in_(box_ids)))
        session.execute(delete(ShipmentBox).where(ShipmentBox.shipment_id.in_(shipment_ids)))
        session.execute(delete(OutboxMessage).where(OutboxMessage.aggregate_id.in_(shipment_ids)))
        session.execute(delete(AuditLog).where(AuditLog.target_id == ORDER_ID))
        session.execute(
            delete(IdempotencyRecord).where(
                IdempotencyRecord.scope.in_(
                    [f"shipment.submit:{user_id}" for user_id in ALL_FACTORY_USER_IDS]
                )
            )
        )
        session.execute(delete(Shipment).where(Shipment.factory_id.in_(FACTORY_IDS)))
        session.execute(
            delete(StoredFile).where(StoredFile.uploaded_by.in_(ALL_FACTORY_USER_IDS))
        )
        session.execute(
            delete(QuantityLedger).where(QuantityLedger.order_assignment_id.in_(assignment_ids))
        )
        session.execute(delete(OrderAssignment).where(OrderAssignment.factory_id.in_(FACTORY_IDS)))
        session.execute(
            delete(OrderCompletionRecord).where(OrderCompletionRecord.order_id == ORDER_ID)
        )
        session.execute(delete(OrderLine).where(OrderLine.order_id == ORDER_ID))
        session.execute(delete(Order).where(Order.order_id == ORDER_ID))
        session.execute(
            delete(UserSession).where(UserSession.user_id.in_([*ALL_FACTORY_USER_IDS, ADMIN_ID]))
        )
        session.execute(delete(ProductVariant).where(ProductVariant.variant_id == VARIANT_ID))
        session.execute(delete(Product).where(Product.product_id == PRODUCT_ID))
        session.execute(delete(User).where(User.user_id.in_([*ALL_FACTORY_USER_IDS, ADMIN_ID])))
        session.execute(delete(Factory).where(Factory.factory_id.in_(FACTORY_IDS)))
        session.execute(
            delete(ShipmentNumberCounter).where(
                ShipmentNumberCounter.business_date == date(2026, 8, 25)
            )
        )


def _seed(engine: Engine, *, initial_shipped_quantity: int = 0) -> int:
    now = datetime(2026, 8, 25, 8, 0)
    with Session(engine) as session, session.begin():
        session.add_all(
            [
                Factory(
                    factory_id=factory_id,
                    supplier_number=f"S07-{index}",
                    factory_name=f"S07接口工厂{index}",
                    factory_code=f"S07-{index}",
                    is_enabled=True,
                )
                for index, factory_id in enumerate(FACTORY_IDS, 1)
            ]
        )
        session.flush()
        session.add_all(
            [
                User(
                    user_id=ADMIN_ID,
                    role="admin",
                    is_enabled=True,
                    feishu_display_name="S08管理员",
                ),
                *[
                    User(
                        user_id=user_id,
                        role="factory",
                        is_enabled=True,
                        feishu_display_name=f"S07工厂用户{index}",
                        factory_id=factory_id,
                        factory_position="employee",
                    )
                    for index, (user_id, factory_id) in enumerate(
                        zip(USER_IDS, FACTORY_IDS, strict=True), 1
                    )
                ],
                User(
                    user_id=SAME_FACTORY_USER_ID,
                    role="factory",
                    is_enabled=True,
                    feishu_display_name="S07同厂用户",
                    factory_id=FACTORY_IDS[0],
                    factory_position="employee",
                ),
            ]
        )
        session.add(
            Product(
                product_id=PRODUCT_ID,
                source_i_id="ITEM-SHIPMENT-API",
                name="S07接口测试产品",
                is_available=True,
                source_modified_at=now,
                first_synced_at=now,
                last_synced_at=now,
            )
        )
        session.add(
            ProductVariant(
                variant_id=VARIANT_ID,
                product_id=PRODUCT_ID,
                source_sku_id="SKU-SHIPMENT-API",
                properties_value="海军蓝 / 120",
                source_category="童帽春夏",
                source_enabled=1,
                is_available=True,
                source_modified_at=now,
                first_synced_at=now,
                last_synced_at=now,
            )
        )
        session.flush()
        session.add(
            Order(
                order_id=ORDER_ID,
                order_no="S07-ORDER-A",
                source="manual",
                order_date=date(2026, 8, 25),
                tracker="松子",
                contract_ship_date=date(2026, 9, 1),
                lifecycle="PUBLISHED",
                version=1,
                published_at=now,
                published_by=USER_IDS[0],
                created_by=USER_IDS[0],
                updated_by=USER_IDS[0],
                created_at=now,
                updated_at=now,
            )
        )
        session.flush()
        line = OrderLine(
            order_id=ORDER_ID,
            product_variant_id=VARIANT_ID,
            order_quantity=40,
            sku_id_snapshot="SKU-SHIPMENT-API",
            product_name_snapshot="S07接口测试产品",
            properties_value_snapshot="海军蓝 / 120",
            category_snapshot="童帽春夏",
            created_at=now,
            updated_at=now,
        )
        session.add(line)
        session.flush()
        assignment = OrderAssignment(
            order_line_id=line.order_line_id,
            factory_id=FACTORY_IDS[0],
            assigned_quantity=40,
            initial_shipped_quantity=initial_shipped_quantity,
            factory_name_snapshot="S07接口工厂1",
            created_at=now,
            updated_at=now,
        )
        session.add(assignment)
        session.flush()
        return assignment.order_assignment_id


def test_factory_creates_one_server_draft_scoped_to_its_factory(
    test_database_engine: Engine,
    test_database_url: str,
) -> None:
    _clean(test_database_engine)
    _seed(test_database_engine)
    sessions = sessionmaker(test_database_engine, class_=Session, expire_on_commit=False)
    identity = IdentityAccessService(
        sessions,
        token_secret=b"shipment-api-token-secret",
        phone_encryption_secret=b"shipment-api-phone-encryption",
        phone_digest_secret=b"shipment-api-phone-digest",
    )
    factory_a = identity.issue_session(user_id=USER_IDS[0], terminal="mini")
    factory_b = identity.issue_session(user_id=USER_IDS[1], terminal="mini")
    app = create_app(database_url=test_database_url, identity_service=identity)

    try:
        with TestClient(app, base_url="https://testserver") as client:
            client.headers["Authorization"] = f"Bearer {factory_a.access_token}"
            created = client.post(
                "/api/v1/factory/shipments/drafts",
                json={"preferredOrderId": ORDER_ID},
            )
            assert created.status_code == 201
            assert {
                key: created.json()[key]
                for key in (
                    "shipmentId",
                    "status",
                    "factoryId",
                    "factoryName",
                    "createdBy",
                    "preferredOrderId",
                )
            } == {
                "shipmentId": created.json()["shipmentId"],
                "status": "DRAFT",
                "factoryId": FACTORY_IDS[0],
                "factoryName": "",
                "createdBy": USER_IDS[0],
                "preferredOrderId": ORDER_ID,
            }

            repeated = client.post(
                "/api/v1/factory/shipments/drafts",
                json={"preferredOrderId": ORDER_ID},
            )
            assert repeated.status_code == 200
            assert repeated.json()["shipmentId"] == created.json()["shipmentId"]

            with TestClient(app, base_url="https://testserver") as other_factory:
                other_factory.headers["Authorization"] = f"Bearer {factory_b.access_token}"
                other_created = other_factory.post(
                    "/api/v1/factory/shipments/drafts",
                    json={},
                )
                assert other_created.status_code == 201
                assert other_created.json()["factoryId"] == FACTORY_IDS[1]
                assert other_created.json()["shipmentId"] != created.json()["shipmentId"]
    finally:
        _clean(test_database_engine)


def test_factory_saves_and_submits_boxes_then_all_terminals_query_same_shipment(
    test_database_engine: Engine,
    test_database_url: str,
) -> None:
    _clean(test_database_engine)
    assignment_id = _seed(test_database_engine, initial_shipped_quantity=5)
    sessions = sessionmaker(test_database_engine, class_=Session, expire_on_commit=False)
    identity = IdentityAccessService(
        sessions,
        token_secret=b"shipment-submit-token-secret",
        phone_encryption_secret=b"shipment-submit-phone-encryption",
        phone_digest_secret=b"shipment-submit-phone-digest",
    )
    factory_a = identity.issue_session(user_id=USER_IDS[0], terminal="mini")
    factory_b = identity.issue_session(user_id=USER_IDS[1], terminal="mini")
    app = create_app(database_url=test_database_url, identity_service=identity)

    try:
        with TestClient(app, base_url="https://testserver") as client:
            client.headers["Authorization"] = f"Bearer {factory_a.access_token}"
            shipment_id = client.post(
                "/api/v1/factory/shipments/drafts",
                json={"preferredOrderId": ORDER_ID},
            ).json()["shipmentId"]
            saved = client.put(
                f"/api/v1/factory/shipments/drafts/{shipment_id}",
                json={
                    "boxes": [
                        {
                            "boxNo": box_no,
                            "groupKey": "group-a" if box_no < 3 else None,
                            "items": [{"assignmentId": assignment_id, "quantity": quantity}],
                        }
                        for box_no, quantity in ((1, 10), (2, 10), (3, 3))
                    ],
                    "note": "尾箱混装",
                },
            )
            assert saved.status_code == 200
            assert saved.json()["totalBoxes"] == 3
            assert saved.json()["totalQuantity"] == 23

            submitted = client.post(
                f"/api/v1/factory/shipments/drafts/{shipment_id}/submit",
                headers={"Idempotency-Key": "shipment-submit-same-key"},
            )
            assert submitted.status_code == 200
            assert submitted.json()["status"] == "SHIPPED"
            assert submitted.json()["shipmentNo"].startswith("FH")
            assert submitted.json()["factoryName"] == "S07接口工厂1"
            assert submitted.json()["lines"][0]["quantity"] == 23

            repeated = client.post(
                f"/api/v1/factory/shipments/drafts/{shipment_id}/submit",
                headers={"Idempotency-Key": "shipment-submit-same-key"},
            )
            assert repeated.json()["shipmentNo"] == submitted.json()["shipmentNo"]

            listing = client.get("/api/v1/factory/shipments")
            assert listing.status_code == 200
            assert listing.json()["total"] == 1
            assert listing.json()["items"][0]["shipmentId"] == shipment_id
            detail = client.get(f"/api/v1/factory/shipments/{shipment_id}")
            assert detail.json()["boxes"][2]["items"][0]["quantity"] == 3
            catalog = client.get("/api/v1/factory/shipment-catalog").json()["items"][0]
            assert catalog["shippedQuantity"] == 28
            assert catalog["pendingQuantity"] == 12
            order = client.get(f"/api/v1/orders/{ORDER_ID}").json()
            assert order["shippedQuantity"] == 28

        with TestClient(app, base_url="https://testserver") as other_factory:
            other_factory.headers["Authorization"] = f"Bearer {factory_b.access_token}"
            assert other_factory.get(f"/api/v1/factory/shipments/{shipment_id}").status_code == 404
    finally:
        _clean(test_database_engine)


def test_factory_uploads_shipment_evidence_and_authorized_terminals_can_read_it(
    test_database_engine: Engine,
    test_database_url: str,
) -> None:
    _clean(test_database_engine)
    assignment_id = _seed(test_database_engine)
    sessions = sessionmaker(test_database_engine, class_=Session, expire_on_commit=False)
    identity = IdentityAccessService(
        sessions,
        token_secret=b"shipment-evidence-token-secret",
        phone_encryption_secret=b"shipment-evidence-phone-encryption",
        phone_digest_secret=b"shipment-evidence-phone-digest",
    )
    factory_a = identity.issue_session(user_id=USER_IDS[0], terminal="mini")
    factory_b = identity.issue_session(user_id=USER_IDS[1], terminal="mini")
    same_factory = identity.issue_session(user_id=SAME_FACTORY_USER_ID, terminal="mini")
    admin_mini = identity.issue_session(user_id=ADMIN_ID, terminal="mini")
    admin_web = identity.issue_session(user_id=ADMIN_ID, terminal="web")
    content = base64.b64decode(
        "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mNk+A8AAQUBAScY42YAAAAASUVORK5CYII="
    )
    private_files = FakePrivateFileStore(bucket="shipment-evidence-test")
    app = create_app(
        database_url=test_database_url,
        identity_service=identity,
        private_file_store=private_files,
    )

    try:
        with TestClient(app, base_url="https://testserver") as factory_client:
            factory_client.headers["Authorization"] = f"Bearer {factory_a.access_token}"
            shipment_id = factory_client.post(
                "/api/v1/factory/shipments/drafts", json={"preferredOrderId": ORDER_ID}
            ).json()["shipmentId"]
            factory_client.put(
                f"/api/v1/factory/shipments/drafts/{shipment_id}",
                json={
                    "boxes": [
                        {
                            "boxNo": 1,
                            "groupKey": None,
                            "items": [{"assignmentId": assignment_id, "quantity": 12}],
                        }
                    ],
                    "note": "附凭证",
                },
            )
            uploaded = factory_client.post(
                f"/api/v1/factory/shipments/drafts/{shipment_id}/files",
                headers={"Idempotency-Key": "shipment-evidence-upload-1"},
                files={"file": ("proof.png", content, "image/png")},
            )
            assert uploaded.status_code == 201
            assert uploaded.json() == {
                "fileId": uploaded.json()["fileId"],
                "filename": "proof.png",
                "mimeType": "image/png",
                "sizeBytes": len(content),
                "contentSha256": hashlib.sha256(content).hexdigest(),
                "displayOrder": 0,
                "contentUrl": f"/api/v1/shipment-files/{uploaded.json()['fileId']}/content",
            }
            repeated = factory_client.post(
                f"/api/v1/factory/shipments/drafts/{shipment_id}/files",
                headers={"Idempotency-Key": "shipment-evidence-upload-1"},
                files={"file": ("proof.png", content, "image/png")},
            )
            assert repeated.status_code == 200
            assert repeated.json()["fileId"] == uploaded.json()["fileId"]
            assert private_files.object_count == 1

            with TestClient(app, base_url="https://testserver") as draft_colleague:
                draft_colleague.headers["Authorization"] = (
                    f"Bearer {same_factory.access_token}"
                )
                assert draft_colleague.get(uploaded.json()["contentUrl"]).content == content
            with TestClient(app, base_url="https://testserver") as draft_admin:
                draft_admin.headers["Authorization"] = f"Bearer {admin_mini.access_token}"
                assert draft_admin.get(uploaded.json()["contentUrl"]).status_code == 404

            submitted = factory_client.post(
                f"/api/v1/factory/shipments/drafts/{shipment_id}/submit",
                headers={"Idempotency-Key": "shipment-evidence-submit"},
            )
            assert submitted.status_code == 200
            assert submitted.json()["files"] == [uploaded.json()]

            own_file = factory_client.get(uploaded.json()["contentUrl"])
            assert own_file.status_code == 200
            assert own_file.content == content
            assert own_file.headers["content-type"] == "image/png"
            assert own_file.headers["cache-control"] == "private, no-store"

        with TestClient(app, base_url="https://testserver") as other_factory:
            other_factory.headers["Authorization"] = f"Bearer {factory_b.access_token}"
            assert other_factory.get(uploaded.json()["contentUrl"]).status_code == 404

        with TestClient(app, base_url="https://testserver") as same_factory_client:
            same_factory_client.headers["Authorization"] = (
                f"Bearer {same_factory.access_token}"
            )
            assert same_factory_client.get(uploaded.json()["contentUrl"]).content == content

        with TestClient(app, base_url="https://testserver") as admin_mini_client:
            admin_mini_client.headers["Authorization"] = f"Bearer {admin_mini.access_token}"
            assert admin_mini_client.get(uploaded.json()["contentUrl"]).content == content

        with TestClient(app, base_url="https://testserver") as admin_web_client:
            admin_web_client.cookies.set("ot_web_session", admin_web.access_token)
            detail = admin_web_client.get(f"/api/v1/admin/shipments/{shipment_id}")
            assert detail.json()["files"] == [uploaded.json()]
            assert admin_web_client.get(uploaded.json()["contentUrl"]).content == content
    finally:
        _clean(test_database_engine)


def test_shipment_evidence_rejects_invalid_or_fourth_image_and_draft_file_can_be_removed(
    test_database_engine: Engine,
    test_database_url: str,
) -> None:
    _clean(test_database_engine)
    assignment_id = _seed(test_database_engine)
    sessions = sessionmaker(test_database_engine, class_=Session, expire_on_commit=False)
    identity = IdentityAccessService(
        sessions,
        token_secret=b"shipment-evidence-validation-token",
        phone_encryption_secret=b"shipment-evidence-validation-phone",
        phone_digest_secret=b"shipment-evidence-validation-digest",
    )
    factory = identity.issue_session(user_id=USER_IDS[0], terminal="mini")
    content = base64.b64decode(
        "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mNk+A8AAQUBAScY42YAAAAASUVORK5CYII="
    )
    private_files = FakePrivateFileStore(bucket="shipment-evidence-validation")
    app = create_app(
        database_url=test_database_url,
        identity_service=identity,
        private_file_store=private_files,
    )

    try:
        with TestClient(app, base_url="https://testserver") as client:
            client.headers["Authorization"] = f"Bearer {factory.access_token}"
            shipment_id = client.post(
                "/api/v1/factory/shipments/drafts", json={"preferredOrderId": ORDER_ID}
            ).json()["shipmentId"]
            client.put(
                f"/api/v1/factory/shipments/drafts/{shipment_id}",
                json={
                    "boxes": [
                        {
                            "boxNo": 1,
                            "groupKey": None,
                            "items": [{"assignmentId": assignment_id, "quantity": 12}],
                        }
                    ],
                    "note": "",
                },
            )
            invalid = client.post(
                f"/api/v1/factory/shipments/drafts/{shipment_id}/files",
                headers={"Idempotency-Key": "shipment-evidence-invalid"},
                files={"file": ("fake.jpg", content, "image/jpeg")},
            )
            assert invalid.status_code == 422
            assert private_files.object_count == 0

            oversized = client.post(
                f"/api/v1/factory/shipments/drafts/{shipment_id}/files",
                headers={"Idempotency-Key": "shipment-evidence-oversized"},
                files={
                    "file": (
                        "oversized.png",
                        content + b"x" * (5 * 1024 * 1024),
                        "image/png",
                    )
                },
            )
            assert oversized.status_code == 422
            assert private_files.object_count == 0

            failing_files = FakePrivateFileStore(
                bucket="shipment-evidence-validation",
                fail_put=True,
            )
            failing_app = create_app(
                database_url=test_database_url,
                identity_service=identity,
                private_file_store=failing_files,
            )
            with TestClient(failing_app, base_url="https://testserver") as failing_client:
                failing_client.headers["Authorization"] = f"Bearer {factory.access_token}"
                failed_upload = failing_client.post(
                    f"/api/v1/factory/shipments/drafts/{shipment_id}/files",
                    headers={"Idempotency-Key": "shipment-evidence-store-failed"},
                    files={"file": ("proof.png", content, "image/png")},
                )
                assert failed_upload.status_code == 503
                assert (
                    failing_client.get("/api/v1/factory/shipments/drafts/current").json()[
                        "files"
                    ]
                    == []
                )

            uploaded = [
                client.post(
                    f"/api/v1/factory/shipments/drafts/{shipment_id}/files",
                    headers={"Idempotency-Key": f"shipment-evidence-{index}"},
                    files={"file": (f"proof-{index}.png", content, "image/png")},
                )
                for index in range(3)
            ]
            assert [item.status_code for item in uploaded] == [201, 201, 201]
            fourth = client.post(
                f"/api/v1/factory/shipments/drafts/{shipment_id}/files",
                headers={"Idempotency-Key": "shipment-evidence-4"},
                files={"file": ("proof-4.png", content, "image/png")},
            )
            assert fourth.status_code == 422
            assert private_files.object_count == 3

            removed = client.delete(
                f"/api/v1/factory/shipments/drafts/{shipment_id}/files/{uploaded[1].json()['fileId']}"
            )
            assert removed.status_code == 204
            draft = client.get("/api/v1/factory/shipments/drafts/current").json()
            assert [item["displayOrder"] for item in draft["files"]] == [0, 1]
            assert [item["fileId"] for item in draft["files"]] == [
                uploaded[0].json()["fileId"],
                uploaded[2].json()["fileId"],
            ]

            submitted = client.post(
                f"/api/v1/factory/shipments/drafts/{shipment_id}/submit",
                headers={"Idempotency-Key": "shipment-evidence-validation-submit"},
            )
            assert len(submitted.json()["files"]) == 2
            cannot_remove = client.delete(
                f"/api/v1/factory/shipments/drafts/{shipment_id}/files/{uploaded[0].json()['fileId']}"
            )
            assert cannot_remove.status_code == 409
    finally:
        _clean(test_database_engine)


def test_factory_catalog_returns_only_its_published_assignments_with_initial_progress(
    test_database_engine: Engine,
    test_database_url: str,
) -> None:
    _clean(test_database_engine)
    assignment_id = _seed(test_database_engine, initial_shipped_quantity=5)
    sessions = sessionmaker(test_database_engine, class_=Session, expire_on_commit=False)
    identity = IdentityAccessService(
        sessions,
        token_secret=b"shipment-catalog-token-secret",
        phone_encryption_secret=b"shipment-catalog-phone-encryption",
        phone_digest_secret=b"shipment-catalog-phone-digest",
    )
    factory_a = identity.issue_session(user_id=USER_IDS[0], terminal="mini")
    factory_b = identity.issue_session(user_id=USER_IDS[1], terminal="mini")
    app = create_app(database_url=test_database_url, identity_service=identity)

    try:
        with TestClient(app, base_url="https://testserver") as client:
            client.headers["Authorization"] = f"Bearer {factory_a.access_token}"
            response = client.get("/api/v1/factory/shipment-catalog")
            assert response.status_code == 200
            assert response.json() == {
                "items": [
                    {
                        "assignmentId": assignment_id,
                        "orderId": ORDER_ID,
                        "orderNo": "S07-ORDER-A",
                        "contractShipDate": "2026-09-01",
                        "productName": "S07接口测试产品",
                        "propertiesValue": "海军蓝 / 120",
                        "assignedQuantity": 40,
                        "shippedQuantity": 5,
                        "pendingQuantity": 35,
                    }
                ],
                "total": 1,
            }

            with TestClient(app, base_url="https://testserver") as other_factory:
                other_factory.headers["Authorization"] = f"Bearer {factory_b.access_token}"
                assert other_factory.get("/api/v1/factory/shipment-catalog").json() == {
                    "items": [],
                    "total": 0,
                }
    finally:
        _clean(test_database_engine)


def test_factory_requests_shipment_withdrawal_without_changing_quantity(
    test_database_engine: Engine,
    test_database_url: str,
) -> None:
    _clean(test_database_engine)
    assignment_id = _seed(test_database_engine)
    sessions = sessionmaker(test_database_engine, class_=Session, expire_on_commit=False)
    identity = IdentityAccessService(
        sessions,
        token_secret=b"shipment-void-token-secret",
        phone_encryption_secret=b"shipment-void-phone-encryption",
        phone_digest_secret=b"shipment-void-phone-digest",
    )
    factory = identity.issue_session(user_id=USER_IDS[0], terminal="mini")
    app = create_app(database_url=test_database_url, identity_service=identity)

    try:
        with TestClient(app, base_url="https://testserver") as client:
            client.headers["Authorization"] = f"Bearer {factory.access_token}"
            shipment_id = client.post(
                "/api/v1/factory/shipments/drafts",
                json={"preferredOrderId": ORDER_ID},
            ).json()["shipmentId"]
            client.put(
                f"/api/v1/factory/shipments/drafts/{shipment_id}",
                json={
                    "boxes": [
                        {
                            "boxNo": 1,
                            "groupKey": None,
                            "items": [{"assignmentId": assignment_id, "quantity": 12}],
                        }
                    ],
                    "note": "",
                },
            )
            client.post(
                f"/api/v1/factory/shipments/drafts/{shipment_id}/submit",
                headers={"Idempotency-Key": "shipment-void-submit"},
            )

            response = client.post(
                f"/api/v1/factory/shipments/{shipment_id}/void-requests",
                headers={"Idempotency-Key": "shipment-void-request"},
                json={"reason": "  装箱数量填写错误  "},
            )

            assert response.status_code == 201
            assert response.json()["shipmentId"] == shipment_id
            assert response.json()["status"] == "PENDING"
            assert response.json()["reason"] == "装箱数量填写错误"
            assert response.json()["requestedByName"] == "S07工厂用户1"
            detail = client.get(f"/api/v1/factory/shipments/{shipment_id}").json()
            assert detail["status"] == "VOID_PENDING"
            assert detail["voidRequest"]["requestId"] == response.json()["requestId"]
            assert detail["voidRequest"]["requestedByName"] == "S07工厂用户1"
            catalog = client.get("/api/v1/factory/shipment-catalog").json()["items"][0]
            assert catalog["shippedQuantity"] == 12
            assert catalog["pendingQuantity"] == 28
    finally:
        _clean(test_database_engine)


def test_admin_approves_withdrawal_once_and_reverses_the_original_quantity(
    test_database_engine: Engine,
    test_database_url: str,
) -> None:
    _clean(test_database_engine)
    assignment_id = _seed(test_database_engine)
    sessions = sessionmaker(test_database_engine, class_=Session, expire_on_commit=False)
    identity = IdentityAccessService(
        sessions,
        token_secret=b"shipment-approve-token-secret",
        phone_encryption_secret=b"shipment-approve-phone-encryption",
        phone_digest_secret=b"shipment-approve-phone-digest",
    )
    factory = identity.issue_session(user_id=USER_IDS[0], terminal="mini")
    admin = identity.issue_session(user_id=ADMIN_ID, terminal="web")
    app = create_app(database_url=test_database_url, identity_service=identity)

    try:
        with TestClient(app, base_url="https://testserver") as factory_client:
            factory_client.headers["Authorization"] = f"Bearer {factory.access_token}"
            shipment_id = factory_client.post(
                "/api/v1/factory/shipments/drafts", json={"preferredOrderId": ORDER_ID}
            ).json()["shipmentId"]
            factory_client.put(
                f"/api/v1/factory/shipments/drafts/{shipment_id}",
                json={
                    "boxes": [
                        {
                            "boxNo": 1,
                            "groupKey": None,
                            "items": [{"assignmentId": assignment_id, "quantity": 12}],
                        }
                    ],
                    "note": "",
                },
            )
            factory_client.post(
                f"/api/v1/factory/shipments/drafts/{shipment_id}/submit",
                headers={"Idempotency-Key": "shipment-approve-submit"},
            )
            with Session(test_database_engine) as session, session.begin():
                order = session.get(Order, ORDER_ID)
                assert order is not None
                order.lifecycle = "COMPLETED"
                order.completed_at = datetime(2026, 8, 25, 9, 0)
                order.completed_by = ADMIN_ID
            request = factory_client.post(
                f"/api/v1/factory/shipments/{shipment_id}/void-requests",
                headers={"Idempotency-Key": "shipment-approve-request"},
                json={"reason": "装箱数量填写错误"},
            ).json()

        with TestClient(app, base_url="https://testserver") as admin_client:
            admin_client.cookies.set("ot_web_session", admin.access_token)
            admin_client.cookies.set("ot_csrf", admin.csrf_token or "")
            response = admin_client.post(
                f"/api/v1/admin/shipment-void-requests/{request['requestId']}/approve",
                headers={
                    "Idempotency-Key": "shipment-approve-review",
                    "X-CSRF-Token": admin.csrf_token or "",
                },
                json={"comment": "确认撤回"},
            )
            assert response.status_code == 200
            assert response.json()["status"] == "APPROVED"
            detail = admin_client.get(f"/api/v1/admin/shipments/{shipment_id}").json()
            assert detail["status"] == "VOIDED"
            audit_contents = [
                item["content"]
                for item in admin_client.get(
                    f"/api/v1/admin/orders/{ORDER_ID}/audit-logs"
                ).json()["items"]
            ]
            assert "通过撤回发货申请，已发数量回退 12 件" in audit_contents
            assert "提交发货单，发货 12 件" in audit_contents
            repeated = admin_client.post(
                f"/api/v1/admin/shipment-void-requests/{request['requestId']}/approve",
                headers={
                    "Idempotency-Key": "shipment-approve-review",
                    "X-CSRF-Token": admin.csrf_token or "",
                },
                json={"comment": "确认撤回"},
            )
            assert repeated.status_code == 200
            assert repeated.json()["status"] == "APPROVED"

        with Session(test_database_engine) as session:
            deltas = list(
                session.scalars(
                    select(QuantityLedger.quantity_delta)
                    .where(QuantityLedger.order_assignment_id == assignment_id)
                    .order_by(QuantityLedger.ledger_id)
                )
            )
            assert deltas == [12, -12]
            order = session.get(Order, ORDER_ID)
            assert order is not None
            assert order.lifecycle == "PUBLISHED"
            reopen = session.scalar(
                select(OrderCompletionRecord)
                .where(OrderCompletionRecord.order_id == ORDER_ID)
                .order_by(OrderCompletionRecord.record_id.desc())
            )
            assert reopen is not None
            assert reopen.action == "REOPEN"
    finally:
        _clean(test_database_engine)


def test_admin_partially_returns_original_shipment_line_and_quantity_can_be_resent(
    test_database_engine: Engine,
    test_database_url: str,
) -> None:
    _clean(test_database_engine)
    assignment_id = _seed(test_database_engine)
    sessions = sessionmaker(test_database_engine, class_=Session, expire_on_commit=False)
    identity = IdentityAccessService(
        sessions,
        token_secret=b"shipment-return-token-secret",
        phone_encryption_secret=b"shipment-return-phone-encryption",
        phone_digest_secret=b"shipment-return-phone-digest",
    )
    factory = identity.issue_session(user_id=USER_IDS[0], terminal="mini")
    admin = identity.issue_session(user_id=ADMIN_ID, terminal="web")
    app = create_app(database_url=test_database_url, identity_service=identity)

    try:
        with TestClient(app, base_url="https://testserver") as factory_client:
            factory_client.headers["Authorization"] = f"Bearer {factory.access_token}"
            shipment_id = factory_client.post(
                "/api/v1/factory/shipments/drafts", json={"preferredOrderId": ORDER_ID}
            ).json()["shipmentId"]
            factory_client.put(
                f"/api/v1/factory/shipments/drafts/{shipment_id}",
                json={
                    "boxes": [
                        {
                            "boxNo": 1,
                            "groupKey": None,
                            "items": [{"assignmentId": assignment_id, "quantity": 12}],
                        }
                    ],
                    "note": "",
                },
            )
            factory_client.post(
                f"/api/v1/factory/shipments/drafts/{shipment_id}/submit",
                headers={"Idempotency-Key": "shipment-return-submit"},
            )
        with Session(test_database_engine) as session:
            shipment_line_id = session.scalar(
                select(ShipmentLine.line_id).where(ShipmentLine.shipment_id == shipment_id)
            )
            assert shipment_line_id is not None

        with TestClient(app, base_url="https://testserver") as admin_client:
            admin_client.cookies.set("ot_web_session", admin.access_token)
            admin_client.cookies.set("ot_csrf", admin.csrf_token or "")
            response = admin_client.post(
                f"/api/v1/admin/shipments/{shipment_id}/returns",
                headers={
                    "Idempotency-Key": "shipment-return-once",
                    "X-CSRF-Token": admin.csrf_token or "",
                },
                json={
                    "reason": "仓库验货退回",
                    "lines": [{"shipmentLineId": shipment_line_id, "quantity": 5}],
                },
            )
            assert response.status_code == 201
            assert response.json()["shipmentId"] == shipment_id
            assert response.json()["lines"][0]["orderNo"] == "S07-ORDER-A"
            assert response.json()["lines"][0]["quantity"] == 5
            repeated = admin_client.post(
                f"/api/v1/admin/shipments/{shipment_id}/returns",
                headers={
                    "Idempotency-Key": "shipment-return-once",
                    "X-CSRF-Token": admin.csrf_token or "",
                },
                json={
                    "reason": "仓库验货退回",
                    "lines": [{"shipmentLineId": shipment_line_id, "quantity": 5}],
                },
            )
            assert repeated.status_code == 200
            assert repeated.json()["eventId"] == response.json()["eventId"]
            detail = admin_client.get(f"/api/v1/admin/shipments/{shipment_id}").json()
            assert detail["lines"][0]["lineId"] == shipment_line_id
            assert detail["lines"][0]["quantity"] == 12
            assert detail["lines"][0]["returnedQuantity"] == 5
            assert detail["lines"][0]["returnableQuantity"] == 7
            over_return = admin_client.post(
                f"/api/v1/admin/shipments/{shipment_id}/returns",
                headers={
                    "Idempotency-Key": "shipment-return-over-limit",
                    "X-CSRF-Token": admin.csrf_token or "",
                },
                json={
                    "reason": "超量退回应失败",
                    "lines": [{"shipmentLineId": shipment_line_id, "quantity": 8}],
                },
            )
            assert over_return.status_code == 409
            decimal_return = admin_client.post(
                f"/api/v1/admin/shipments/{shipment_id}/returns",
                headers={
                    "Idempotency-Key": "shipment-return-decimal",
                    "X-CSRF-Token": admin.csrf_token or "",
                },
                json={
                    "reason": "小数退回应失败",
                    "lines": [{"shipmentLineId": shipment_line_id, "quantity": 1.5}],
                },
            )
            assert decimal_return.status_code == 422

        with Session(test_database_engine) as session:
            deltas = list(
                session.scalars(
                    select(QuantityLedger.quantity_delta)
                    .where(QuantityLedger.order_assignment_id == assignment_id)
                    .order_by(QuantityLedger.ledger_id)
                )
            )
            assert deltas == [12, -5]
    finally:
        _clean(test_database_engine)


def test_returned_shipment_may_request_withdrawal_but_can_only_be_rejected(
    test_database_engine: Engine,
    test_database_url: str,
) -> None:
    _clean(test_database_engine)
    assignment_id = _seed(test_database_engine)
    sessions = sessionmaker(test_database_engine, class_=Session, expire_on_commit=False)
    identity = IdentityAccessService(
        sessions,
        token_secret=b"shipment-return-void-token",
        phone_encryption_secret=b"shipment-return-void-phone",
        phone_digest_secret=b"shipment-return-void-digest",
    )
    factory = identity.issue_session(user_id=USER_IDS[0], terminal="mini")
    admin = identity.issue_session(user_id=ADMIN_ID, terminal="web")
    app = create_app(database_url=test_database_url, identity_service=identity)

    try:
        with TestClient(app, base_url="https://testserver") as factory_client:
            factory_client.headers["Authorization"] = f"Bearer {factory.access_token}"
            shipment_id = factory_client.post(
                "/api/v1/factory/shipments/drafts", json={"preferredOrderId": ORDER_ID}
            ).json()["shipmentId"]
            factory_client.put(
                f"/api/v1/factory/shipments/drafts/{shipment_id}",
                json={
                    "boxes": [
                        {
                            "boxNo": 1,
                            "groupKey": None,
                            "items": [{"assignmentId": assignment_id, "quantity": 12}],
                        }
                    ],
                    "note": "",
                },
            )
            factory_client.post(
                f"/api/v1/factory/shipments/drafts/{shipment_id}/submit",
                headers={"Idempotency-Key": "shipment-return-void-submit"},
            )
        with Session(test_database_engine) as session:
            shipment_line_id = session.scalar(
                select(ShipmentLine.line_id).where(ShipmentLine.shipment_id == shipment_id)
            )
            assert shipment_line_id is not None

        with TestClient(app, base_url="https://testserver") as admin_client:
            admin_client.cookies.set("ot_web_session", admin.access_token)
            admin_client.cookies.set("ot_csrf", admin.csrf_token or "")
            write_headers = {"X-CSRF-Token": admin.csrf_token or ""}
            returned = admin_client.post(
                f"/api/v1/admin/shipments/{shipment_id}/returns",
                headers={**write_headers, "Idempotency-Key": "return-before-void"},
                json={
                    "reason": "部分退回",
                    "lines": [{"shipmentLineId": shipment_line_id, "quantity": 3}],
                },
            )
            assert returned.status_code == 201

        with TestClient(app, base_url="https://testserver") as factory_client:
            factory_client.headers["Authorization"] = f"Bearer {factory.access_token}"
            request = factory_client.post(
                f"/api/v1/factory/shipments/{shipment_id}/void-requests",
                headers={"Idempotency-Key": "void-after-return"},
                json={"reason": "仍需申请撤回"},
            )
            assert request.status_code == 201

        with TestClient(app, base_url="https://testserver") as admin_client:
            admin_client.cookies.set("ot_web_session", admin.access_token)
            admin_client.cookies.set("ot_csrf", admin.csrf_token or "")
            write_headers = {"X-CSRF-Token": admin.csrf_token or ""}
            blocked_return = admin_client.post(
                f"/api/v1/admin/shipments/{shipment_id}/returns",
                headers={**write_headers, "Idempotency-Key": "return-while-void-pending"},
                json={
                    "reason": "待审期间不可退回",
                    "lines": [{"shipmentLineId": shipment_line_id, "quantity": 1}],
                },
            )
            assert blocked_return.status_code == 409
            blocked_approve = admin_client.post(
                f"/api/v1/admin/shipment-void-requests/{request.json()['requestId']}/approve",
                headers={**write_headers, "Idempotency-Key": "approve-returned-shipment"},
                json={"comment": ""},
            )
            assert blocked_approve.status_code == 409
            rejected = admin_client.post(
                f"/api/v1/admin/shipment-void-requests/{request.json()['requestId']}/reject",
                headers={**write_headers, "Idempotency-Key": "reject-returned-shipment"},
                json={"comment": "已有退回，只能拒绝"},
            )
            assert rejected.status_code == 200
            detail = admin_client.get(f"/api/v1/admin/shipments/{shipment_id}").json()
            assert detail["status"] == "SHIPPED"
            assert detail["lines"][0]["returnableQuantity"] == 9

        with Session(test_database_engine) as session:
            deltas = list(
                session.scalars(
                    select(QuantityLedger.quantity_delta)
                    .where(QuantityLedger.order_assignment_id == assignment_id)
                    .order_by(QuantityLedger.ledger_id)
                )
            )
            assert deltas == [12, -3]
    finally:
        _clean(test_database_engine)


def test_web_admin_downloads_shipment_workbook_from_submitted_facts(
    test_database_engine: Engine,
    test_database_url: str,
) -> None:
    _clean(test_database_engine)
    assignment_id = _seed(test_database_engine)
    sessions = sessionmaker(test_database_engine, class_=Session, expire_on_commit=False)
    identity = IdentityAccessService(
        sessions,
        token_secret=b"shipment-export-token",
        phone_encryption_secret=b"shipment-export-phone",
        phone_digest_secret=b"shipment-export-digest",
    )
    factory = identity.issue_session(user_id=USER_IDS[0], terminal="mini")
    admin = identity.issue_session(user_id=ADMIN_ID, terminal="web")
    app = create_app(database_url=test_database_url, identity_service=identity)

    try:
        with TestClient(app, base_url="https://testserver") as factory_client:
            factory_client.headers["Authorization"] = f"Bearer {factory.access_token}"
            shipment_id = factory_client.post(
                "/api/v1/factory/shipments/drafts", json={"preferredOrderId": ORDER_ID}
            ).json()["shipmentId"]
            factory_client.put(
                f"/api/v1/factory/shipments/drafts/{shipment_id}",
                json={
                    "boxes": [
                        {
                            "boxNo": 1,
                            "groupKey": None,
                            "items": [{"assignmentId": assignment_id, "quantity": 12}],
                        }
                    ],
                    "note": "",
                },
            )
            submitted = factory_client.post(
                f"/api/v1/factory/shipments/drafts/{shipment_id}/submit",
                headers={"Idempotency-Key": "shipment-export-submit"},
            ).json()

        with TestClient(app, base_url="https://testserver") as admin_client:
            admin_client.cookies.set("ot_web_session", admin.access_token)
            response = admin_client.get(f"/api/v1/admin/shipments/{shipment_id}/export")
            assert response.status_code == 200
            assert response.headers["content-type"] == (
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            assert "filename*=UTF-8''" in response.headers["content-disposition"]
            workbook = load_workbook(BytesIO(response.content), data_only=False)
            assert workbook.sheetnames == ["发货明细", "汇总"]
            assert workbook["发货明细"]["A3"].value == "S07-ORDER-A"
            assert workbook["发货明细"]["F3"].value == 12
            assert workbook["汇总"]["D2"].value == 12
            assert submitted["shipmentNo"] in response.headers["content-disposition"]
    finally:
        _clean(test_database_engine)


def test_factory_can_save_and_reopen_empty_and_partly_packed_boxes(
    test_database_engine: Engine,
    test_database_url: str,
) -> None:
    assignment_id = _seed(test_database_engine)
    sessions = sessionmaker(
        test_database_engine, class_=Session, expire_on_commit=False
    )
    identity = IdentityAccessService(
        sessions,
        token_secret=b"draft-test",
        phone_encryption_secret=b"draft-phone",
        phone_digest_secret=b"draft-digest",
    )
    token = identity.issue_session(user_id=USER_IDS[0], terminal="mini")
    app = create_app(database_url=test_database_url, identity_service=identity)
    headers = {"Authorization": f"Bearer {token.access_token}"}
    with TestClient(app, base_url="https://testserver", headers=headers) as client:
        draft = client.post("/api/v1/factory/shipments/drafts", json={}).json()
        path = f"/api/v1/factory/shipments/drafts/{draft['shipmentId']}"
        saved = client.put(
            path,
            json={
                "boxes": [
                    {"boxNo": 1, "items": []},
                    {"boxNo": 2, "items": []},
                ],
                "note": "半成品草稿",
            },
        )
        assert saved.status_code == 200, saved.text
        assert saved.json()["totalBoxes"] == 2
        boxes = [
            {"boxNo": 1, "items": [{"assignmentId": assignment_id, "quantity": 5}]},
            {"boxNo": 2, "items": []},
        ]
        saved = client.put(
            path,
            json={
                "boxes": boxes,
                "note": "先装一箱",
                "version": saved.json()["version"],
            },
        )
        assert saved.status_code == 200, saved.text
        assert (
            client.post(
                path + "/submit", headers={"Idempotency-Key": "empty-box"}
            ).status_code
            == 422
        )
    with TestClient(app, base_url="https://testserver", headers=headers) as reopened:
        current = reopened.get("/api/v1/factory/shipments/drafts/current").json()
        assert current["shipmentId"] == draft["shipmentId"]
        assert current["totalBoxes"] == 2
        assert current["totalQuantity"] == 5
        assert current["boxes"][1]["items"] == []
        assert current["note"] == "先装一箱"


def test_factory_can_abandon_only_own_current_version_and_restart(
    test_database_engine: Engine,
    test_database_url: str,
) -> None:
    _seed(test_database_engine)
    sessions = sessionmaker(
        test_database_engine, class_=Session, expire_on_commit=False
    )
    identity = IdentityAccessService(
        sessions,
        token_secret=b"draft-delete",
        phone_encryption_secret=b"draft-phone",
        phone_digest_secret=b"draft-digest",
    )
    app = create_app(database_url=test_database_url, identity_service=identity)
    owner = identity.issue_session(user_id=USER_IDS[0], terminal="mini")
    colleague = identity.issue_session(user_id=SAME_FACTORY_USER_ID, terminal="mini")
    with TestClient(app, base_url="https://testserver") as client:
        client.headers["Authorization"] = f"Bearer {owner.access_token}"
        draft = client.post("/api/v1/factory/shipments/drafts", json={}).json()
        path = f"/api/v1/factory/shipments/drafts/{draft['shipmentId']}"
        saved = client.put(
            path,
            json={
                "boxes": [{"boxNo": 1, "items": []}],
                "note": "保留新内容",
                "version": 1,
            },
        ).json()
        assert (
            client.put(
                path,
                json={
                    "boxes": [{"boxNo": 1, "items": []}],
                    "note": "过时覆盖",
                    "version": 1,
                },
            ).status_code
            == 409
        )
        client.headers["Authorization"] = f"Bearer {colleague.access_token}"
        assert client.get("/api/v1/factory/shipments/drafts/current").status_code == 404
        assert (
            client.delete(path, params={"version": saved["version"]}).status_code == 404
        )
        client.headers["Authorization"] = f"Bearer {owner.access_token}"
        assert client.delete(path, params={"version": 1}).status_code == 409
        assert (
            client.get("/api/v1/factory/shipments/drafts/current").json()["note"]
            == "保留新内容"
        )
        assert (
            client.delete(path, params={"version": saved["version"]}).status_code == 204
        )
        assert client.get("/api/v1/factory/shipments/drafts/current").status_code == 404
        assert (
            client.put(path, json={"boxes": [{"boxNo": 1, "items": []}]}).status_code
            == 404
        )
        replacement = client.post("/api/v1/factory/shipments/drafts", json={}).json()
        assert replacement["shipmentId"] != draft["shipmentId"]
        assert client.get("/api/v1/factory/shipments").json()["items"] == []


def test_concurrent_draft_creation_and_saves_do_not_duplicate_or_overwrite(
    test_database_engine: Engine,
    test_database_url: str,
) -> None:
    from concurrent.futures import ThreadPoolExecutor
    from threading import Barrier

    assignment_id = _seed(test_database_engine)
    sessions = sessionmaker(
        test_database_engine, class_=Session, expire_on_commit=False
    )
    identity = IdentityAccessService(
        sessions,
        token_secret=b"draft-race",
        phone_encryption_secret=b"draft-phone",
        phone_digest_secret=b"draft-digest",
    )
    token = identity.issue_session(user_id=USER_IDS[0], terminal="mini")
    app = create_app(database_url=test_database_url, identity_service=identity)
    headers = {"Authorization": f"Bearer {token.access_token}"}
    gate = Barrier(2)

    def create() -> dict:
        with TestClient(app, base_url="https://testserver", headers=headers) as client:
            gate.wait()
            response = client.post("/api/v1/factory/shipments/drafts", json={})
            assert response.status_code in (200, 201), response.text
            return response.json()

    with ThreadPoolExecutor(max_workers=2) as pool:
        a, b = list(pool.map(lambda _: create(), range(2)))
    assert a["shipmentId"] == b["shipmentId"]
    path = f"/api/v1/factory/shipments/drafts/{a['shipmentId']}"
    boxes = [{"boxNo": 1, "items": [{"assignmentId": assignment_id, "quantity": 3}]}]

    def save(note: str) -> tuple[int, str]:
        with TestClient(app, base_url="https://testserver", headers=headers) as client:
            gate.wait()
            response = client.put(
                path, json={"version": 1, "boxes": boxes, "note": note}
            )
            return response.status_code, note

    with ThreadPoolExecutor(max_workers=2) as pool:
        results = list(pool.map(save, ["设备甲", "设备乙"]))
    assert sorted(code for code, _ in results) == [200, 409]
    winning_note = next(note for code, note in results if code == 200)
    with TestClient(app, base_url="https://testserver", headers=headers) as client:
        current = client.get("/api/v1/factory/shipments/drafts/current").json()
        assert current["note"] == winning_note
        assert (
            client.post(
                path + "/submit?version=1", headers={"Idempotency-Key": "stale"}
            ).status_code
            == 409
        )
        submitted = client.post(
            path + f"/submit?version={current['version']}",
            headers={"Idempotency-Key": "fresh"},
        )
        assert submitted.status_code == 200
        repeated = client.post(
            path + f"/submit?version={current['version']}",
            headers={"Idempotency-Key": "retry"},
        )
        assert repeated.json()["shipmentNo"] == submitted.json()["shipmentNo"]
        assert repeated.json()["totalQuantity"] == 3
        assert client.get("/api/v1/factory/shipments/drafts/current").status_code == 404
        assert (
            client.delete(path, params={"version": current["version"]}).status_code
            == 409
        )


def test_retrying_acknowledged_save_preserves_draft_and_uploaded_evidence(
    test_database_engine: Engine,
    test_database_url: str,
) -> None:
    _seed(test_database_engine)
    sessions = sessionmaker(
        test_database_engine, class_=Session, expire_on_commit=False
    )
    identity = IdentityAccessService(
        sessions,
        token_secret=b"draft-evidence",
        phone_encryption_secret=b"draft-phone",
        phone_digest_secret=b"draft-digest",
    )
    token = identity.issue_session(user_id=USER_IDS[0], terminal="mini")
    app = create_app(
        database_url=test_database_url,
        identity_service=identity,
        private_file_store=FakePrivateFileStore(bucket="draft-evidence-test"),
    )
    content = base64.b64decode(
        "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mNk+A8AAQUBAScY42YAAAAASUVORK5CYII="
    )
    with TestClient(app, base_url="https://testserver") as client:
        client.headers["Authorization"] = f"Bearer {token.access_token}"
        draft = client.post("/api/v1/factory/shipments/drafts", json={}).json()
        path = f"/api/v1/factory/shipments/drafts/{draft['shipmentId']}"
        payload = {
            "version": 1,
            "boxes": [{"boxNo": 1, "items": []}],
            "note": "保存成功回包丢失",
        }
        saved = client.put(path, json=payload).json()
        uploaded = client.post(
            path + "/files",
            headers={"Idempotency-Key": "proof"},
            files={"file": ("proof.png", content, "image/png")},
        ).json()
        retried = client.put(path, json=payload)
        assert retried.status_code == 200, retried.text
        assert retried.json()["version"] == saved["version"]
        current = client.get("/api/v1/factory/shipments/drafts/current").json()
        assert current["files"] == [uploaded]
        assert current["note"] == payload["note"]
        assert client.get(uploaded["contentUrl"]).content == content
        reused = client.post("/api/v1/factory/shipments/drafts", json={}).json()
        assert reused == current
        assert (
            client.delete(path, params={"version": current["version"]}).status_code
            == 204
        )
        assert client.get(uploaded["contentUrl"]).status_code == 404


def test_admin_filters_shipments_by_actual_order_lines(
    test_database_engine: Engine, test_database_url: str
) -> None:
    assignment_id = _seed(test_database_engine, initial_shipped_quantity=5)
    with Session(test_database_engine) as session, session.begin():
        session.add(
            Order(
                order_id="second-order",
                order_no="OTHER-ORDER",
                source="manual",
                tracker="松子",
                contract_ship_date=date(2026, 9, 30),
                lifecycle="PUBLISHED",
                created_by=ADMIN_ID,
                updated_by=ADMIN_ID,
            )
        )
        session.flush()
        line = OrderLine(
            order_id="second-order",
            product_variant_id=VARIANT_ID,
            order_quantity=40,
            sku_id_snapshot="SKU-SHIPMENT-API",
            product_name_snapshot="第二订单",
            properties_value_snapshot="海军蓝 / 120",
        )
        session.add(line)
        session.flush()
        other = OrderAssignment(
            order_line_id=line.order_line_id,
            factory_id=FACTORY_IDS[0],
            assigned_quantity=40,
            factory_name_snapshot="S07接口工厂1",
        )
        session.add(other)
        session.flush()
        other_assignment_id = other.order_assignment_id
    sessions = sessionmaker(test_database_engine, class_=Session, expire_on_commit=False)
    identity = IdentityAccessService(
        sessions,
        token_secret=b"filter-token",
        phone_encryption_secret=b"filter-phone",
        phone_digest_secret=b"filter-digest",
    )
    admin = identity.issue_session(user_id=ADMIN_ID, terminal="web")
    factory = identity.issue_session(user_id=USER_IDS[0], terminal="mini")
    app = create_app(database_url=test_database_url, identity_service=identity)
    with TestClient(app, base_url="https://testserver") as client:
        client.cookies.set("ot_web_session", admin.access_token)
        assert (
            client.get("/api/v1/admin/shipments", params={"orderId": ORDER_ID}).json()["items"]
            == []
        )
        with TestClient(app, base_url="https://testserver") as factory_client:
            factory_client.headers["Authorization"] = f"Bearer {factory.access_token}"
            assert (
                factory_client.get(
                    "/api/v1/admin/shipments", params={"orderId": ORDER_ID}
                ).status_code
                == 403
            )
            shipment_ids = []
            for index in range(2):
                shipment_id = factory_client.post(
                    "/api/v1/factory/shipments/drafts", json={}
                ).json()["shipmentId"]
                assert (
                    factory_client.put(
                        f"/api/v1/factory/shipments/drafts/{shipment_id}",
                        json={
                            "boxes": [
                                {
                                    "boxNo": 1,
                                    "items": [
                                        {"assignmentId": assignment_id, "quantity": 2},
                                        *(
                                            [{"assignmentId": other_assignment_id, "quantity": 3}]
                                            if index == 0
                                            else []
                                        ),
                                    ],
                                }
                            ]
                        },
                    ).status_code
                    == 200
                )
                assert (
                    factory_client.post(
                        f"/api/v1/factory/shipments/drafts/{shipment_id}/submit",
                        headers={"Idempotency-Key": f"filter-{index}"},
                    ).status_code
                    == 200
                )
                shipment_ids.append(shipment_id)
        result = client.get("/api/v1/admin/shipments", params={"orderId": ORDER_ID}).json()
        assert result["total"] == 2
        assert {item["shipmentId"] for item in result["items"]} == set(shipment_ids)
        assert all(item["preferredOrderId"] is None for item in result["items"])
        second = client.get("/api/v1/admin/shipments", params={"orderId": "second-order"}).json()
        assert [item["shipmentId"] for item in second["items"]] == [shipment_ids[0]]
        assert (
            next(item for item in result["items"] if item["shipmentId"] == shipment_ids[0])[
                "totalQuantity"
            ]
            == 5
        )
        with Session(test_database_engine) as session, session.begin():
            session.get(Shipment, shipment_ids[0]).status = "VOIDED"
            session.get(Shipment, shipment_ids[1]).status = "VOID_PENDING"
        statuses = client.get("/api/v1/admin/shipments", params={"orderId": ORDER_ID}).json()
        assert {item["status"] for item in statuses["items"]} == {"VOIDED", "VOID_PENDING"}
        unrelated = client.get(
            "/api/v1/admin/shipments", params={"orderId": "unrelated-order"}
        ).json()
        assert unrelated["items"] == []
        assert client.get("/api/v1/admin/shipments").json()["total"] == 2
