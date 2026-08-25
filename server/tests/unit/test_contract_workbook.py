from io import BytesIO
from pathlib import Path
from zipfile import ZipFile

import pytest
from openpyxl import load_workbook
from PIL import Image

from app.modules.contracts.workbook import ContractWorkbookRenderer


def _snapshot(lines: list[dict[str, object]]) -> dict[str, object]:
    return {
        "contractNo": "20260824-KK-HT",
        "signingDate": "2026-08-24",
        "orderNo": "HT-ORDER-BOUNDARY",
        "contractShipDate": "2026-09-10",
        "factory": {
            "legalName": "合同测试工厂有限公司",
            "address": "浙江省杭州市测试路1号",
            "legalRepresentative": "测试法人",
            "phone": "",
        },
        "lines": lines,
    }


@pytest.mark.parametrize("line_count", [1, 12])
def test_renderer_supports_original_detail_area_boundaries(line_count: int) -> None:
    template = Path(__file__).resolve().parents[2] / "app/templates/processing_contract_v1.xlsx"
    renderer = ContractWorkbookRenderer(template_path=template)
    lines = [
        {
            "productId": "product-1",
            "itemNo": "MZ2026-01",
            "productName": "儿童遮阳帽",
            "propertiesValue": f"规格 {index + 1}",
            "quantity": index + 1,
            "imageObjectKey": None,
        }
        for index in range(line_count)
    ]

    content = renderer.render(_snapshot(lines))

    sheet = load_workbook(BytesIO(content), data_only=False)["合同"]
    last_row = 7 + line_count
    assert sheet.cell(last_row, 4).value == f"规格 {line_count}"
    assert sheet["E20"].value == f"=SUM(E8:E{last_row})"
    assert str(sheet.print_area) == "'合同'!$A$1:$I$51"
    with ZipFile(BytesIO(content)) as archive:
        unsafe = ("vbaProject", "externalLink", "oleObject", "embeddings")
        assert not any(any(marker in name for marker in unsafe) for name in archive.namelist())


def test_renderer_keeps_template_layout_and_leaves_prices_and_incomplete_totals_blank() -> None:
    template = Path(__file__).resolve().parents[2] / "app/templates/processing_contract_v1.xlsx"
    renderer = ContractWorkbookRenderer(template_path=template)

    content = renderer.render(
        {
            "contractNo": "20260824-KK-HT",
            "signingDate": "2026-08-24",
            "orderNo": "HT-ORDER-001",
            "contractShipDate": "2026-09-10",
            "factory": {
                "legalName": "合同测试工厂有限公司",
                "address": "浙江省杭州市测试路1号",
                "legalRepresentative": "测试法人",
                "phone": "13800000000",
            },
            "lines": [
                {
                    "productId": "product-1",
                    "itemNo": "MZ2026-01",
                    "productName": "儿童遮阳帽",
                    "propertiesValue": "米色 / 52cm",
                    "quantity": 40,
                    "imageObjectKey": None,
                },
                {
                    "productId": "product-1",
                    "itemNo": "MZ2026-01",
                    "productName": "儿童遮阳帽",
                    "propertiesValue": "米色 / 54cm",
                    "quantity": 60,
                    "imageObjectKey": None,
                },
            ],
        }
    )

    workbook = load_workbook(BytesIO(content), data_only=False)
    sheet = workbook["合同"]
    assert workbook.sheetnames == ["合同"]
    assert sheet["G3"].value == "合同编号：20260824-KK-HT"
    assert sheet["G4"].value == "签订时间：2026年8月24日"
    assert sheet["A4"].value == "供方：合同测试工厂有限公司"
    assert sheet["A8"].value == "MZ2026-01"
    assert sheet["B8"].value == "儿童遮阳帽"
    assert sheet["D8"].value == "米色 / 52cm"
    assert sheet["D9"].value == "米色 / 54cm"
    assert sheet["E8"].value == 40
    assert sheet["E9"].value == 60
    assert sheet["F8"].value is None
    assert sheet["F9"].value is None
    assert sheet["G8"].value == '=IF(OR(E8="",F8=""),"",E8*F8)'
    assert sheet["G9"].value == '=IF(OR(E9="",F9=""),"",E9*F9)'
    assert sheet["G10"].value is None
    assert sheet["E20"].value == "=SUM(E8:E9)"
    assert sheet["G20"].value == '=IF(COUNT(F8:F9)=ROWS(F8:F9),SUM(G8:G9),"")'
    assert sheet["B21"].value == '=IF(G20="","",G20)'
    assert sheet["D21"].value == '=IF(G20="","",G20)'
    assert "A8:A9" in {str(item) for item in sheet.merged_cells.ranges}
    assert "B8:B9" in {str(item) for item in sheet.merged_cells.ranges}
    assert "C8:C9" in {str(item) for item in sheet.merged_cells.ranges}
    assert "H8:H19" in {str(item) for item in sheet.merged_cells.ranges}
    assert sheet["H8"].value == "2026年9月10日"
    assert "合同测试工厂有限公司" in sheet["E45"].value
    assert "委托代理人：\n开户银行：\n账号：\n电话：13800000000" in sheet["E45"].value
    assert all(
        cell.fill.fill_type is None
        for row in sheet.iter_rows(min_row=1, max_row=51, min_col=1, max_col=9)
        for cell in row
    )
    assert "南昌昱斌" not in "".join(
        str(cell.value or "") for row in sheet.iter_rows() for cell in row
    )


def test_renderer_expands_contract_beyond_twelve_detail_rows() -> None:
    template = Path(__file__).resolve().parents[2] / "app/templates/processing_contract_v1.xlsx"
    renderer = ContractWorkbookRenderer(template_path=template)
    lines = [
        {
            "productId": "product-1",
            "itemNo": "MZ2026-01",
            "productName": "儿童遮阳帽",
            "propertiesValue": f"规格 {index + 1}",
            "quantity": index + 1,
            "imageObjectKey": None,
        }
        for index in range(13)
    ]

    content = renderer.render(
        {
            "contractNo": "20260824-KK-HT-1",
            "signingDate": "2026-08-24",
            "orderNo": "HT-ORDER-002",
            "contractShipDate": "2026-09-10",
            "factory": {
                "legalName": "合同测试工厂有限公司",
                "address": "浙江省杭州市测试路1号",
                "legalRepresentative": "测试法人",
                "phone": "",
            },
            "lines": lines,
        }
    )

    workbook = load_workbook(BytesIO(content), data_only=False)
    sheet = workbook["合同"]
    assert sheet["D20"].value == "规格 13"
    assert sheet["G20"].value == '=IF(OR(E20="",F20=""),"",E20*F20)'
    assert sheet["E21"].value == "=SUM(E8:E20)"
    assert sheet["G21"].value == '=IF(COUNT(F8:F20)=ROWS(F8:F20),SUM(G8:G20),"")'
    assert "A8:A20" in {str(item) for item in sheet.merged_cells.ranges}
    assert "合同测试工厂有限公司" in sheet["E46"].value
    assert str(sheet.print_area) == "'合同'!$A$1:$I$52"


def test_renderer_embeds_available_product_image_once_for_product_group() -> None:
    template = Path(__file__).resolve().parents[2] / "app/templates/processing_contract_v1.xlsx"
    image_bytes = BytesIO()
    Image.new("RGB", (200, 100), color=(20, 100, 70)).save(image_bytes, format="PNG")
    renderer = ContractWorkbookRenderer(
        template_path=template,
        image_loader=lambda object_key: image_bytes.getvalue()
        if object_key == "products/hat.png"
        else None,
    )

    content = renderer.render(
        {
            "contractNo": "20260824-KK-HT-2",
            "signingDate": "2026-08-24",
            "orderNo": "HT-ORDER-003",
            "contractShipDate": "2026-09-10",
            "factory": {
                "legalName": "合同测试工厂有限公司",
                "address": "浙江省杭州市测试路1号",
                "legalRepresentative": "测试法人",
                "phone": "",
            },
            "lines": [
                {
                    "productId": "product-1",
                    "itemNo": "MZ2026-01",
                    "productName": "儿童遮阳帽",
                    "propertiesValue": "米色 / 52cm",
                    "quantity": 40,
                    "imageObjectKey": "products/hat.png",
                },
                {
                    "productId": "product-1",
                    "itemNo": "MZ2026-01",
                    "productName": "儿童遮阳帽",
                    "propertiesValue": "米色 / 54cm",
                    "quantity": 60,
                    "imageObjectKey": "products/hat.png",
                },
            ],
        }
    )

    workbook = load_workbook(BytesIO(content))
    sheet = workbook["合同"]
    assert len(sheet._images) == 1
    image = sheet._images[0]
    assert image.anchor._from.col == 2
    assert image.anchor._from.row == 7


def test_renderer_preserves_table_styles_for_each_product_group() -> None:
    template = Path(__file__).resolve().parents[2] / "app/templates/processing_contract_v1.xlsx"
    renderer = ContractWorkbookRenderer(template_path=template)
    lines = [
        {
            "productId": f"product-{index}",
            "itemNo": f"ITEM-{index}",
            "productName": f"产品 {index}",
            "propertiesValue": f"规格 {index}",
            "quantity": index,
            "imageObjectKey": None,
        }
        for index in (1, 2)
    ]

    sheet = load_workbook(BytesIO(renderer.render(_snapshot(lines))))["合同"]

    for row in (8, 9):
        for column in (1, 2, 3):
            cell = sheet.cell(row, column)
            assert cell.style_id != 0
            assert cell.border.left.style == "thin"
            assert cell.border.right.style == "thin"
            assert cell.border.bottom.style == "thin"
