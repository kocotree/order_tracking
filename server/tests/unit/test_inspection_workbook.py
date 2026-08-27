from io import BytesIO
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile

import pytest
from openpyxl import load_workbook

from app.modules.repairs.workbook import (
    InspectionWorkbookLimits,
    InspectionWorkbookParser,
    InspectionWorkbookValidationError,
)


def test_parser_normalizes_basic_inspection_workbook_in_original_row_order() -> None:
    source = Path(__file__).resolve().parents[3] / "docs/reference/E28质检.xlsx"

    snapshot = InspectionWorkbookParser().parse(source.read_bytes())

    assert snapshot.supplier_number == "E28"
    assert snapshot.factory_name == "跃富"
    assert snapshot.total_quantity == 126
    assert snapshot.box_numbers == ("1号箱",)
    assert len(snapshot.lines) == 10
    assert snapshot.lines[0].source_row == 2
    assert snapshot.lines[0].box_number == "1号箱"
    assert snapshot.lines[0].source_sku_id == "6941716599133"
    assert snapshot.lines[0].source_product_id == "KQ26022"
    assert snapshot.lines[0].product_name == "小动物软檐鸭舌帽"
    assert snapshot.lines[0].properties_value == "兔兔奶糖S"
    assert snapshot.lines[0].quantity == 51
    assert snapshot.lines[0].reason is None
    assert snapshot.lines[-1].source_row == 11
    assert snapshot.lines[-1].box_number == "1号箱"
    assert snapshot.lines[-1].source_sku_id == "6942649403252"
    assert snapshot.lines[-1].quantity == 57


def test_parser_expands_reason_by_its_own_merge_range() -> None:
    source = Path(__file__).resolve().parents[3] / "docs/reference/E22质检.xlsx"

    snapshot = InspectionWorkbookParser().parse(source.read_bytes())

    assert snapshot.supplier_number == "E22"
    assert snapshot.factory_name == "宇婷"
    assert snapshot.total_quantity == 826
    assert snapshot.box_numbers == ("1号箱", "2号箱", "3号箱")
    assert len(snapshot.lines) == 33
    assert snapshot.lines[0].reason == "脏   面料次\n破   针眼"
    assert snapshot.lines[20].source_row == 22
    assert snapshot.lines[20].box_number == "2号箱"
    assert snapshot.lines[20].reason == "脏   面料次\n破   针眼"
    assert snapshot.lines[21].source_row == 23
    assert snapshot.lines[21].box_number == "3号箱"
    assert snapshot.lines[21].reason == "线   面料次\n脏   缝次"
    assert snapshot.lines[-1].reason == "线   面料次\n脏   缝次"


def test_parser_ignores_photo_cells_and_returns_only_structured_repair_fields() -> None:
    source = Path(__file__).resolve().parents[3] / "docs/reference/E22质检.xlsx"
    workbook = load_workbook(source)
    workbook["Sheet1"]["J2"] = "=SUM(1,1)"
    modified = BytesIO()
    workbook.save(modified)

    snapshot = InspectionWorkbookParser().parse(modified.getvalue())

    assert snapshot.total_quantity == 826
    assert snapshot.lines[0].reason == "脏   面料次\n破   针眼"
    assert not hasattr(snapshot, "image_count")
    assert not hasattr(snapshot.lines[0], "images")


def test_parser_reports_invalid_header_at_sheet_row_and_field_boundary() -> None:
    source = Path(__file__).resolve().parents[3] / "docs/reference/E28质检.xlsx"
    workbook = load_workbook(source)
    workbook["Sheet1"]["A1"] = "错误编号"
    modified = BytesIO()
    workbook.save(modified)

    with pytest.raises(InspectionWorkbookValidationError) as caught:
        InspectionWorkbookParser().parse(modified.getvalue())

    assert caught.value.issues == (
        {
            "code": "invalid_header",
            "message": "A 列表头应为“编号”",
            "sheet": "Sheet1",
            "row": 1,
            "field": "A",
        },
    )


def test_parser_reports_non_positive_quantity_at_source_row() -> None:
    source = Path(__file__).resolve().parents[3] / "docs/reference/E28质检.xlsx"
    workbook = load_workbook(source)
    workbook["Sheet1"]["G2"] = 0
    modified = BytesIO()
    workbook.save(modified)

    with pytest.raises(InspectionWorkbookValidationError) as caught:
        InspectionWorkbookParser().parse(modified.getvalue())

    assert caught.value.issues == (
        {
            "code": "invalid_quantity",
            "message": "仓库退回数量必须是正整数",
            "sheet": "Sheet1",
            "row": 2,
            "field": "G",
        },
    )


def test_parser_rejects_source_before_opening_ooxml_when_file_limit_is_exceeded() -> None:
    source = Path(__file__).resolve().parents[3] / "docs/reference/E28质检.xlsx"
    parser = InspectionWorkbookParser(limits=InspectionWorkbookLimits(max_source_bytes=1))

    with pytest.raises(InspectionWorkbookValidationError) as caught:
        parser.parse(source.read_bytes())

    assert caught.value.issues == (
        {
            "code": "file_too_large",
            "message": "质检 Excel 超过文件大小上限",
        },
    )


def test_parser_rejects_ooxml_when_zip_entry_limit_is_exceeded() -> None:
    source = Path(__file__).resolve().parents[3] / "docs/reference/E28质检.xlsx"
    parser = InspectionWorkbookParser(limits=InspectionWorkbookLimits(max_zip_entries=1))

    with pytest.raises(InspectionWorkbookValidationError) as caught:
        parser.parse(source.read_bytes())

    assert caught.value.issues == (
        {
            "code": "too_many_zip_entries",
            "message": "质检 Excel 内部文件数超过上限",
        },
    )


def test_parser_rejects_ooxml_when_uncompressed_total_limit_is_exceeded() -> None:
    source = Path(__file__).resolve().parents[3] / "docs/reference/E28质检.xlsx"
    parser = InspectionWorkbookParser(
        limits=InspectionWorkbookLimits(max_uncompressed_bytes=1)
    )

    with pytest.raises(InspectionWorkbookValidationError) as caught:
        parser.parse(source.read_bytes())

    assert caught.value.issues == (
        {
            "code": "uncompressed_content_too_large",
            "message": "质检 Excel 解压后总量超过上限",
        },
    )


def test_parser_rejects_external_ooxml_relationships() -> None:
    source = Path(__file__).resolve().parents[3] / "docs/reference/E28质检.xlsx"
    modified = BytesIO()
    with ZipFile(BytesIO(source.read_bytes())) as original, ZipFile(
        modified, "w", ZIP_DEFLATED
    ) as rewritten:
        for entry in original.infolist():
            data = original.read(entry.filename)
            if entry.filename == "xl/_rels/workbook.xml.rels":
                data = data.replace(
                    b"</Relationships>",
                    b'<Relationship Id="external-test" Target="https://example.invalid/file" '
                    b'TargetMode="External" Type="external-test"/></Relationships>',
                )
            rewritten.writestr(entry, data)

    with pytest.raises(InspectionWorkbookValidationError) as caught:
        InspectionWorkbookParser().parse(modified.getvalue())

    assert caught.value.issues == (
        {
            "code": "external_relationship",
            "message": "质检 Excel 不允许引用外部资源",
        },
    )


def test_parser_rejects_unsafe_zip_entry_paths() -> None:
    source = Path(__file__).resolve().parents[3] / "docs/reference/E28质检.xlsx"
    modified = BytesIO()
    with ZipFile(BytesIO(source.read_bytes())) as original, ZipFile(
        modified, "w", ZIP_DEFLATED
    ) as rewritten:
        for entry in original.infolist():
            rewritten.writestr(entry, original.read(entry.filename))
        rewritten.writestr("../escape.txt", b"unsafe")

    with pytest.raises(InspectionWorkbookValidationError) as caught:
        InspectionWorkbookParser().parse(modified.getvalue())

    assert caught.value.issues == (
        {
            "code": "unsafe_archive_path",
            "message": "质检 Excel 包含不安全的内部路径",
        },
    )


def test_parser_rejects_workbook_when_worksheet_limit_is_exceeded() -> None:
    source = Path(__file__).resolve().parents[3] / "docs/reference/E28质检.xlsx"
    parser = InspectionWorkbookParser(
        limits=InspectionWorkbookLimits(max_worksheets=2)
    )

    with pytest.raises(InspectionWorkbookValidationError) as caught:
        parser.parse(source.read_bytes())

    assert caught.value.issues == (
        {
            "code": "too_many_worksheets",
            "message": "质检 Excel 工作表数量超过上限",
        },
    )


def test_parser_rejects_workbook_when_data_row_limit_is_exceeded() -> None:
    source = Path(__file__).resolve().parents[3] / "docs/reference/E28质检.xlsx"
    parser = InspectionWorkbookParser(limits=InspectionWorkbookLimits(max_data_rows=9))

    with pytest.raises(InspectionWorkbookValidationError) as caught:
        parser.parse(source.read_bytes())

    assert caught.value.issues == (
        {
            "code": "too_many_data_rows",
            "message": "质检 Excel 数据行数超过上限",
            "sheet": "Sheet1",
        },
    )


def test_parser_reports_missing_required_base_field_at_source_row() -> None:
    source = Path(__file__).resolve().parents[3] / "docs/reference/E28质检.xlsx"
    workbook = load_workbook(source)
    workbook["Sheet1"]["C3"] = None
    modified = BytesIO()
    workbook.save(modified)

    with pytest.raises(InspectionWorkbookValidationError) as caught:
        InspectionWorkbookParser().parse(modified.getvalue())

    assert caught.value.issues == (
        {
            "code": "required_field_missing",
            "message": "商品编码不能为空",
            "sheet": "Sheet1",
            "row": 3,
            "field": "C",
        },
    )


def test_parser_reports_missing_box_number_at_source_row() -> None:
    source = Path(__file__).resolve().parents[3] / "docs/reference/E28质检.xlsx"
    workbook = load_workbook(source)
    workbook["Sheet1"].unmerge_cells("H2:H11")
    workbook["Sheet1"]["H2"] = None
    modified = BytesIO()
    workbook.save(modified)

    with pytest.raises(InspectionWorkbookValidationError) as caught:
        InspectionWorkbookParser().parse(modified.getvalue())

    assert caught.value.issues == (
        {
            "code": "required_field_missing",
            "message": "箱数不能为空",
            "sheet": "Sheet1",
            "row": 2,
            "field": "H",
        },
    )


def test_parser_rejects_business_data_after_first_blank_base_row() -> None:
    source = Path(__file__).resolve().parents[3] / "docs/reference/E28质检.xlsx"
    workbook = load_workbook(source)
    for column in range(1, 8):
        workbook["Sheet1"].cell(5, column).value = None
    modified = BytesIO()
    workbook.save(modified)

    with pytest.raises(InspectionWorkbookValidationError) as caught:
        InspectionWorkbookParser().parse(modified.getvalue())

    assert caught.value.issues == (
        {
            "code": "data_after_blank_row",
            "message": "空白数据行之后不能再出现业务数据",
            "sheet": "Sheet1",
            "row": 6,
        },
    )


def test_parser_rejects_business_content_outside_sheet1() -> None:
    source = Path(__file__).resolve().parents[3] / "docs/reference/E28质检.xlsx"
    workbook = load_workbook(source)
    workbook["Sheet2"]["A1"] = "不应忽略的业务数据"
    modified = BytesIO()
    workbook.save(modified)

    with pytest.raises(InspectionWorkbookValidationError) as caught:
        InspectionWorkbookParser().parse(modified.getvalue())

    assert caught.value.issues == (
        {
            "code": "unexpected_sheet_data",
            "message": "Sheet1 以外的工作表不能包含业务数据",
            "sheet": "Sheet2",
            "row": 1,
            "field": "A",
        },
    )
