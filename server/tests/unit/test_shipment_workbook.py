from datetime import date
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

from app.modules.shipments.workbook import (
    ShipmentWorkbookLine,
    ShipmentWorkbookRenderer,
    ShipmentWorkbookSnapshot,
)


def test_shipment_workbook_keeps_detail_orders_and_merges_summary_across_orders() -> None:
    template = Path(__file__).resolve().parents[3] / "docs/reference/厂家发货模版.xlsx"
    renderer = ShipmentWorkbookRenderer(template_path=template)

    content = renderer.render(
        ShipmentWorkbookSnapshot(
            business_date=date(2026, 8, 25),
            total_boxes=2,
            lines=[
                ShipmentWorkbookLine(
                    order_no="ORDER-A",
                    box_no="1",
                    sku_id="SKU-001",
                    product_name="儿童遮阳帽",
                    properties_value="米白 / 52cm",
                    packed_quantity=6,
                    total_quantity=6,
                ),
                ShipmentWorkbookLine(
                    order_no="ORDER-B",
                    box_no="2",
                    sku_id="SKU-001",
                    product_name="儿童遮阳帽",
                    properties_value="米白 / 52cm",
                    packed_quantity=4,
                    total_quantity=4,
                ),
            ],
        )
    )

    workbook = load_workbook(BytesIO(content), data_only=False)
    assert workbook.sheetnames == ["发货明细", "汇总"]
    detail = workbook["发货明细"]
    assert detail["A1"].value == "KK发货清单 2026年8月25日 共计2箱"
    assert [detail.cell(2, column).value for column in range(1, 8)] == [
        "订单编号",
        "箱号",
        "货号",
        "品名",
        "颜色/规格",
        "装箱数量",
        "合计",
    ]
    assert [detail.cell(3, column).value for column in range(1, 8)] == [
        "ORDER-A",
        "1",
        "SKU-001",
        "儿童遮阳帽",
        "米白 / 52cm",
        6,
        6,
    ]
    assert detail["A4"].value == "ORDER-B"

    summary = workbook["汇总"]
    assert [summary.cell(1, column).value for column in range(1, 5)] == [
        "日期",
        "名称",
        "颜色/规格",
        "数量",
    ]
    assert summary["A2"].value.date() == date(2026, 8, 25)
    assert [summary.cell(2, column).value for column in range(2, 5)] == [
        "儿童遮阳帽",
        "米白 / 52cm",
        10,
    ]
    assert summary.max_row == 2
